import re
import math
from datetime import datetime
from pathlib import Path

import altair as alt
import pandas as pd
import streamlit as st
from streamlit.errors import StreamlitAPIException


DATA_FILE = "깃허브_S관(3공장) 계획대시보드.csv"
FX_FILE = "월별업데이트_환율기준.csv"
WORKDAYS_FILE = "근무일수(기준자료).csv"
TARGET_FILE = "월별업데이트_생산목표량.csv"
GOAL_SUMMARY_FILE = "깃허브_S관공장목표현황_요약.csv"
FAST_MODE_KEY = "fast_mode"
MAX_STYLE_ROWS = 1500
SPEC_NUMBER_RE = re.compile(r"[+-]\d+\.\d{2}")
CYL_AXIS_RE = re.compile(r"([+-]\d+\.\d{2})(\d{3})")
DEBUG_DUPLICATES = False
PACK_COL_CANDIDATES = ["포장단위", "포장단위명", "포장규격", "단위", "UOM", "uom", "PACK_UOM", "포장 UOM"]
SAMPLE_EXCL_COL_CANDIDATES = [
    "샘플제외 양품수량",
    "샘플제외양품수량",
    "샘플제외_양품수량",
    "샘플제외 양품",
]
PROCESS_LABELS = {
    "[10]": "사출조립",
    "[20]": "분리",
    "[45]": "수화/검사",
    "[80]": "누수/규격검사",
}
PROCESS_KEYS = list(PROCESS_LABELS.keys())


def format_date_series(series):
    dates = pd.to_datetime(series, errors="coerce")
    return dates.dt.strftime("%Y-%m-%d").fillna("")


def parse_spec_from_code(code):
    if not code:
        return "", "", ""
    text = str(code)
    numbers = SPEC_NUMBER_RE.findall(text)
    power = numbers[0] if numbers else ""
    cylinder = numbers[1] if len(numbers) > 1 else ""
    axis = ""
    cyl_match = CYL_AXIS_RE.search(text)
    if cyl_match:
        axis = cyl_match.group(2)
        if not cylinder:
            cylinder = cyl_match.group(1)
    return power, cylinder, axis


def find_first_column(columns, candidates):
    for name in candidates:
        if name in columns:
            return name
    return None


def normalize_numeric(series):
    return (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.strip()
        .replace({"": "0", "nan": "0", "None": "0"})
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0)
    )


def normalize_process_code(value):
    text = str(value).strip()
    if not text:
        return ""
    for key in PROCESS_KEYS:
        if key in text:
            return key
    digits = re.findall(r"\d+", text)
    if digits:
        code = digits[0]
        if code in {"10", "20", "45", "80"}:
            return f"[{code}]"
    return text


def process_display(code):
    return PROCESS_LABELS.get(code, code)


def debug_duplicate_orders(view_df, label):
    base_cols = [col for col in ["이니셜", "품목코드", "수주번호"] if col in view_df.columns]
    pack_col = find_first_column(view_df.columns, PACK_COL_CANDIDATES)
    key_cols = base_cols + [col for col in [pack_col] if col]
    if not key_cols:
        print(f"[DEBUG] {label}: no key columns available.")
        return
    key_df = view_df[key_cols].astype(str).fillna("")
    counts = key_df.value_counts()
    dupes = counts[counts > 1]
    if dupes.empty:
        print(f"[DEBUG] {label}: no duplicate keys found.")
        return
    print(f"[DEBUG] {label}: duplicate keys {len(dupes)}")
    for key, count in dupes.head(20).items():
        print(f"[DEBUG] {key} -> {count}")


def add_spec_columns(view):
    def pick_code(row):
        for col in ("품목코드", "Q코드", "R코드"):
            value = row.get(col, "")
            if pd.notna(value):
                text = str(value).strip()
                if text:
                    return text
        return ""

    specs = view.apply(
        lambda row: parse_spec_from_code(pick_code(row)),
        axis=1,
        result_type="expand",
    )
    view["파워"] = specs[0]
    view["실린더(ADD)"] = specs[1]
    view["축"] = specs[2]


def join_unique(series):
    values = sorted(
        {str(value).strip() for value in series if pd.notna(value) and str(value).strip()}
    )
    return ";".join(values) if values else ""


def normalize_name_tokens(text):
    cleaned = re.sub(r"\([^)]*\)", "", str(text))
    cleaned = cleaned.replace(";", " ")
    tokens = []
    for token in cleaned.split():
        token = token.strip(" ,;")
        if not token:
            continue
        if "_" in token:
            parts = token.split("_")
            if any(char.isdigit() for char in parts[-1]):
                token = parts[0]
        tokens.append(token)
    return tokens


def representative_name(series):
    names = [str(value).strip() for value in series if pd.notna(value) and str(value).strip()]
    unique = list(dict.fromkeys(names))
    if not unique:
        return ""
    if len(unique) == 1:
        return unique[0]
    token_lists = [normalize_name_tokens(name) for name in unique]
    base_tokens = token_lists[0]
    if not base_tokens:
        return unique[0]
    token_sets = [set(tokens) for tokens in token_lists[1:]]
    common = [token for token in base_tokens if all(token in s for s in token_sets)]
    if len(common) >= 2:
        return " ".join(common).strip()
    return unique[0]


def fill_object_na(frame):
    df = frame.copy()
    object_cols = df.select_dtypes(include=["object"]).columns
    if len(object_cols) > 0:
        df[object_cols] = df[object_cols].fillna("")
    return df


def apply_alignment(styler, frame):
    numeric_cols = [
        col for col in frame.columns if pd.api.types.is_numeric_dtype(frame[col])
    ]
    text_cols = [col for col in frame.columns if col not in numeric_cols]
    if text_cols:
        styler = styler.set_properties(subset=text_cols, **{"text-align": "center"})
    if numeric_cols:
        styler = styler.set_properties(subset=numeric_cols, **{"text-align": "right"})
    styler = styler.set_table_styles(
        [{"selector": "th", "props": [("text-align", "center")]}],
        overwrite=False,
    )
    return styler


def render_dataframe(styled, fallback, **kwargs):
    fast_mode = st.session_state.get(FAST_MODE_KEY, False)
    if fast_mode:
        return st.dataframe(fallback, **kwargs)
    if isinstance(fallback, pd.DataFrame) and len(fallback) > MAX_STYLE_ROWS:
        return st.dataframe(fallback, **kwargs)
    try:
        return st.dataframe(styled, **kwargs)
    except StreamlitAPIException:
        return st.dataframe(fallback, **kwargs)


def calc_table_height(
    row_count, row_height=32, header_height=40, min_height=140, max_height=650
):
    if row_count <= 0:
        return min_height
    height = header_height + row_height * (row_count + 1)
    return max(min_height, min(max_height, int(height)))


def render_selection_sum(df, selected_rows, label="선택 합계"):
    if not selected_rows:
        return
    numeric_cols = df.select_dtypes(include="number").columns
    if numeric_cols.empty:
        return
    sums = df.iloc[selected_rows][numeric_cols].sum()
    sum_df = sums.to_frame().T
    sum_df.insert(0, "구분", label)
    format_dict = {col: "{:,.0f}" for col in numeric_cols}
    st.caption(label)
    summary_styled = apply_alignment(
        sum_df.style.format(format_dict, na_rep=""),
        sum_df,
    )
    st.dataframe(
        summary_styled,
        width="stretch",
        height=calc_table_height(1, min_height=120, max_height=180),
    )


def add_due_warning(frame, date_col="출고예상일", days=7):
    if date_col not in frame.columns:
        return frame, None, None
    df = frame.copy()
    dates = pd.to_datetime(df[date_col], errors="coerce")
    today = pd.Timestamp.today().normalize()
    overdue = dates < today
    soon = dates.between(today, today + pd.Timedelta(days=days))
    date_str = dates.dt.strftime("%Y-%m-%d").fillna("")
    date_str = date_str.where(~(overdue | soon), "❗ " + date_str)
    df[date_col] = date_str
    return df, overdue, soon


def aggregate_production(view):
    if "품목코드" not in view.columns:
        return view

    pack_col = find_first_column(view.columns, PACK_COL_CANDIDATES)
    key_cols = [col for col in ["이니셜", "품목코드", "수주번호"] if col in view.columns]
    if pack_col:
        key_cols.append(pack_col)
    if key_cols:
        view = view.drop_duplicates(subset=key_cols)

    stock_cols = [
        "사출창고",
        "분리창고",
        "검사접착",
        "누수규격",
        "완제품",
        "불용재고",
    ]
    sum_cols = [
        "수량",
        "잔여수주량",
        "생산필요량",
        "사출필요량",
        "분리필요량",
        "수화필요량",
        "접착필요량",
        "누수/규격필요량",
    ]
    for col in sum_cols + stock_cols:
        if col in view.columns:
            view[col] = pd.to_numeric(view[col], errors="coerce").fillna(0)

    stock_cols_available = [col for col in stock_cols if col in view.columns]
    stock_data = view.drop_duplicates(subset=["품목코드"])[
        ["품목코드"] + stock_cols_available
    ]

    agg = {}
    for col in sum_cols:
        if col in view.columns:
            agg[col] = "sum"
    if "출고예상일" in view.columns:
        agg["출고예상일"] = "min"
    for col in ["품명", "생산품명", "신규분류코드", "Q코드", "R코드"]:
        if col in view.columns:
            agg[col] = representative_name if col == "품명" else join_unique

    view_without_stock = view.drop(columns=stock_cols_available, errors="ignore")
    grouped = view_without_stock.groupby("품목코드", dropna=False).agg(agg).reset_index()
    if stock_cols_available:
        grouped = grouped.merge(stock_data, on="품목코드", how="left")
    return grouped


def filter_need_rows(frame):
    need_cols = [col for col in ("생산필요량", "사출필요량", "분리필요량") if col in frame.columns]
    if not need_cols:
        return frame
    mask = pd.Series(False, index=frame.index)
    for col in need_cols:
        mask |= pd.to_numeric(frame[col], errors="coerce").fillna(0) > 0
    return frame[mask]


def load_data(path):
    last_error = None
    for enc in ("utf-8-sig", "cp949"):
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    raise last_error


def load_fx(path):
    last_error = None
    for enc in ("utf-8-sig", "cp949"):
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    raise last_error


@st.cache_data(show_spinner=False)
def load_actuals():
    files = [
        path
        for path in Path(".").glob("일별업데이트_생산실적현황*.csv")
        if "2025" in path.name or "2026" in path.name
    ]
    frames = []
    for path in files:
        try:
            df = load_data(path)
        except Exception:
            continue
        df.columns = [c.strip() for c in df.columns]
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


@st.cache_data(show_spinner=False)
def load_targets():
    path = Path(TARGET_FILE)
    if not path.exists():
        return pd.DataFrame()
    df = load_data(path)
    df.columns = [c.strip() for c in df.columns]
    return df


@st.cache_data(show_spinner=False)
def load_workdays():
    path = Path(WORKDAYS_FILE)
    if not path.exists():
        return pd.DataFrame()
    df = load_data(path)
    df.columns = [c.strip() for c in df.columns]
    return df


@st.cache_data(show_spinner=False)
def load_goal_summary():
    path = Path(GOAL_SUMMARY_FILE)
    if not path.exists():
        return pd.DataFrame()
    df = load_data(path)
    df.columns = [c.strip() for c in df.columns]
    if "일자" in df.columns:
        df["일자"] = pd.to_datetime(df["일자"], errors="coerce")
    if "공정코드" in df.columns:
        df["공정코드"] = df["공정코드"].apply(normalize_process_code)
    for col in [
        "실적",
        "실적_양품",
        "실적_샘플제외",
        "생산수량",
        "불량수량",
        "샘플수량",
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            df[col] = 0
    if "연" not in df.columns and "일자" in df.columns:
        df["연"] = df["일자"].dt.year
    if "월" not in df.columns and "일자" in df.columns:
        df["월"] = df["일자"].dt.month
    if "실제근무" not in df.columns:
        df["실제근무"] = (df["실적_양품"] > 0) | (df["생산수량"] > 0)
    return df


def aggregate_actuals_from_daily(daily_df):
    if daily_df.empty:
        return pd.DataFrame(
            columns=[
                "연",
                "월",
                "공정코드",
                "실적",
                "실적_양품",
                "실적_샘플제외",
                "생산수량",
                "불량수량",
                "샘플수량",
                "실제근무일수",
                "수율",
            ]
        )
    workdays = (
        daily_df[daily_df["실제근무"]]
        .groupby(["연", "월", "공정코드"], dropna=False)["일자"]
        .nunique()
        .reset_index(name="실제근무일수")
    )
    grouped = (
        daily_df.groupby(["연", "월", "공정코드"], dropna=False)
        .agg(
            실적_양품=("실적_양품", "sum"),
            실적_샘플제외=("실적_샘플제외", "sum"),
            생산수량=("생산수량", "sum"),
            불량수량=("불량수량", "sum"),
            샘플수량=("샘플수량", "sum"),
        )
        .reset_index()
    )
    grouped = grouped.merge(workdays, on=["연", "월", "공정코드"], how="left")
    grouped["실제근무일수"] = grouped["실제근무일수"].fillna(0)
    grouped["실적"] = grouped["실적_양품"]
    grouped["수율"] = grouped["실적"] / grouped["생산수량"].replace(0, pd.NA)
    return grouped


def preprocess_actuals(actuals):
    if actuals.empty:
        return pd.DataFrame(
            columns=[
                "연",
                "월",
                "공정코드",
                "실적",
                "실적_양품",
                "실적_샘플제외",
                "생산수량",
                "불량수량",
                "샘플수량",
                "실제근무일수",
                "수율",
            ]
        )
    df = actuals.copy()
    df["공장"] = df["공장"].astype(str).str.strip()
    df = df[df["공장"] == "S관(3공장)"]
    df["공정코드"] = df["공정코드"].apply(normalize_process_code)
    df = df[df["공정코드"].isin(PROCESS_KEYS)]
    df["생산일자"] = pd.to_datetime(df["생산일자"], errors="coerce")
    df = df[df["생산일자"].dt.year.isin([2025, 2026])]
    df["연"] = df["생산일자"].dt.year
    df["월"] = df["생산일자"].dt.month
    df["양품수량"] = normalize_numeric(df["양품수량"])
    sample_excl_col = find_first_column(df.columns, SAMPLE_EXCL_COL_CANDIDATES)
    if sample_excl_col:
        df[sample_excl_col] = normalize_numeric(df[sample_excl_col])
    else:
        df["샘플제외양품수량"] = 0
        sample_excl_col = "샘플제외양품수량"
    for col in ["생산수량", "불량수량", "샘플수량"]:
        if col in df.columns:
            df[col] = normalize_numeric(df[col])
        else:
            df[col] = 0
    df["실제근무"] = (df["양품수량"] > 0) | (df["생산수량"] > 0)

    workdays = (
        df[df["실제근무"]]
        .groupby(["연", "월", "공정코드"], dropna=False)["생산일자"]
        .nunique()
        .reset_index(name="실제근무일수")
    )
    grouped = (
        df.groupby(["연", "월", "공정코드"], dropna=False)
        .agg(
            실적_양품=("양품수량", "sum"),
            실적_샘플제외=(sample_excl_col, "sum"),
            생산수량=("생산수량", "sum"),
            불량수량=("불량수량", "sum"),
            샘플수량=("샘플수량", "sum"),
        )
        .reset_index()
    )
    grouped = grouped.merge(workdays, on=["연", "월", "공정코드"], how="left")
    grouped["실적"] = grouped["실적_양품"]
    grouped["실제근무일수"] = grouped["실제근무일수"].fillna(0)
    grouped["수율"] = grouped["실적"] / grouped["생산수량"].replace(0, pd.NA)
    return grouped


def preprocess_targets(targets):
    if targets.empty:
        return pd.DataFrame(columns=["연", "월", "공정코드", "일일_생산목표량"])
    df = targets.copy()
    df.columns = [c.strip() for c in df.columns]
    target_col = find_first_column(
        df.columns,
        [
            "일일_생산목표량",
            "일일 생산목표량",
            "일일_생산목표량",
            "일일_생산목표량",
        ],
    )
    if target_col is None:
        return pd.DataFrame(columns=["연", "월", "공정코드", "일일_생산목표량"])
    df["년"] = pd.to_numeric(df["년"], errors="coerce")
    df["월"] = pd.to_numeric(df["월"], errors="coerce")
    df["공장"] = df["공장"].astype(str).str.strip()
    df["공정코드"] = df["공정코드"].apply(normalize_process_code)
    df = df[df["공장"] == "S관(3공장)"]
    df = df[df["공정코드"].isin(PROCESS_KEYS)]
    df = df[df["년"].isin([2025, 2026])]
    df[target_col] = (
        df[target_col].astype(str).str.replace(",", "").str.strip()
    )
    df[target_col] = pd.to_numeric(df[target_col], errors="coerce").fillna(0)
    grouped = (
        df.groupby(["년", "월", "공정코드"], dropna=False)[target_col]
        .mean()
        .reset_index()
        .rename(columns={"년": "연", target_col: "일일_생산목표량"})
    )
    return grouped


def preprocess_workdays(workdays):
    if workdays.empty or "합계" not in workdays.columns:
        return pd.DataFrame(columns=["연", "월", "기준자료근무일수", "기준근무일수"])
    df = workdays.copy()
    df["합계"] = df["합계"].astype(str).str.strip()
    melted = df.melt(id_vars=["합계"], var_name="월", value_name="값")
    melted["월"] = (
        melted["월"].astype(str).str.replace("월", "").str.strip()
    )
    melted["월"] = pd.to_numeric(melted["월"], errors="coerce")
    melted["값"] = (
        melted["값"].astype(str).str.replace("일", "").str.replace(",", "").str.strip()
    )
    melted["값"] = pd.to_numeric(melted["값"], errors="coerce").fillna(0)

    base = melted[melted["합계"] == "기준일"][["월", "값"]].rename(
        columns={"값": "기준근무일수"}
    )
    yearly = melted[melted["합계"].str.contains("년")].copy()
    yearly["연"] = yearly["합계"].str.replace("년", "").str.strip()
    yearly["연"] = pd.to_numeric(yearly["연"], errors="coerce")
    yearly = yearly.rename(columns={"값": "기준자료근무일수"})
    merged = yearly.merge(base, on="월", how="left")
    return merged[["연", "월", "기준자료근무일수", "기준근무일수"]]


def preprocess_actuals_daily(actuals):
    if actuals.empty:
        return pd.DataFrame(
            columns=[
                "일자",
                "연",
                "월",
                "공정코드",
                "실적",
                "실적_양품",
                "실적_샘플제외",
                "생산수량",
                "불량수량",
                "샘플수량",
                "수율",
                "실제근무",
            ]
        )
    df = actuals.copy()
    df["공장"] = df["공장"].astype(str).str.strip()
    df = df[df["공장"] == "S관(3공장)"]
    df["공정코드"] = df["공정코드"].apply(normalize_process_code)
    df = df[df["공정코드"].isin(PROCESS_KEYS)]
    df["생산일자"] = pd.to_datetime(df["생산일자"], errors="coerce")
    df = df[df["생산일자"].dt.year.isin([2025, 2026])]
    df["일자"] = df["생산일자"].dt.normalize()
    df["연"] = df["생산일자"].dt.year
    df["월"] = df["생산일자"].dt.month
    df["양품수량"] = normalize_numeric(df["양품수량"])
    sample_excl_col = find_first_column(df.columns, SAMPLE_EXCL_COL_CANDIDATES)
    if sample_excl_col:
        df[sample_excl_col] = normalize_numeric(df[sample_excl_col])
    else:
        df["샘플제외양품수량"] = 0
        sample_excl_col = "샘플제외양품수량"
    for col in ["생산수량", "불량수량", "샘플수량"]:
        if col in df.columns:
            df[col] = normalize_numeric(df[col])
        else:
            df[col] = 0
    df["실제근무"] = (df["양품수량"] > 0) | (df["생산수량"] > 0)
    grouped = (
        df.groupby(["일자", "연", "월", "공정코드"], dropna=False)
        .agg(
            실적_양품=("양품수량", "sum"),
            실적_샘플제외=(sample_excl_col, "sum"),
            생산수량=("생산수량", "sum"),
            불량수량=("불량수량", "sum"),
            샘플수량=("샘플수량", "sum"),
            실제근무=("실제근무", "max"),
        )
        .reset_index()
    )
    grouped["실적"] = grouped["실적_양품"]
    grouped["수율"] = grouped["실적"] / grouped["생산수량"].replace(0, pd.NA)
    return grouped


def compute_daily_table(
    daily_df,
    targets_df,
    workdays_df,
    selected_date,
    compare_date,
    process_codes,
    target_mode,
    view_mode,
    actual_col="실적",
):
    if daily_df.empty:
        return pd.DataFrame()
    actual_col = actual_col if actual_col in daily_df.columns else "실적"
    good_col = "실적_양품" if "실적_양품" in daily_df.columns else actual_col
    target_lookup = (
        targets_df[(targets_df["연"] == selected_date.year) & (targets_df["월"] == selected_date.month)]
        .set_index("공정코드")["일일_생산목표량"]
        .to_dict()
    )
    workdays_month = workdays_df[
        (workdays_df["연"] == selected_date.year) & (workdays_df["월"] == selected_date.month)
    ]
    workdays_value = (
        workdays_month["기준자료근무일수"].iloc[0] if not workdays_month.empty else 0
    )

    def actual_slice(start_date, end_date):
        return daily_df[
            (daily_df["일자"] >= start_date)
            & (daily_df["일자"] <= end_date)
            & (daily_df["공정코드"].isin(process_codes))
        ]


    if view_mode == "Daily":
        data_today = actual_slice(selected_date, selected_date)
        data_prev = actual_slice(compare_date, compare_date)
        rows = []
        for code in process_codes:
            today_row = data_today[data_today["공정코드"] == code]
            prev_row = data_prev[data_prev["공정코드"] == code]
            actual_today = today_row[actual_col].sum()
            actual_prev = prev_row[actual_col].sum()
            prod_today = today_row["생산수량"].sum()
            prod_prev = prev_row["생산수량"].sum()
            bad_today = today_row["불량수량"].sum()
            sample_today = today_row["샘플수량"].sum()
            total_out_today = (
                today_row[good_col].sum() + bad_today + sample_today
            )
            total_out_prev = (
                prev_row[good_col].sum()
                + prev_row["불량수량"].sum()
                + prev_row["샘플수량"].sum()
            )
            denom_today = max(prod_today, total_out_today) if total_out_today else prod_today
            denom_prev = max(prod_prev, total_out_prev) if total_out_prev else prod_prev
            yield_today = actual_today / prod_today if prod_today else pd.NA
            yield_adj_today = actual_today / denom_today if denom_today else pd.NA
            yield_prev = actual_prev / prod_prev if prod_prev else pd.NA
            target_daily = target_lookup.get(code, 0)
            rate_today = actual_today / target_daily if target_daily else pd.NA
            rate_prev = actual_prev / target_daily if target_daily else pd.NA
            anomaly = total_out_today > prod_today
            rows.append(
                {
                    "공정": process_display(code),
                    "목표": target_daily,
                    "실적(선택일)": actual_today,
                    "실적(전일)": actual_prev,
                    "실적 증감": actual_today - actual_prev,
                    "달성율(선택일)": rate_today,
                    "달성율(전일)": rate_prev,
                    "달성율 증감": (
                        rate_today - rate_prev if pd.notna(rate_today) and pd.notna(rate_prev) else pd.NA
                    ),
                    "생산수량": prod_today,
                    "수율(원칙)": yield_today,
                    "수율(보정)": yield_adj_today,
                    "불량수량": bad_today,
                    "샘플수량": sample_today,
                    "수율이상여부": "Y" if anomaly else "",
                    "이상원인": "분모불일치" if anomaly else "",
                }
            )
        return pd.DataFrame(rows)

    start_date = selected_date.replace(day=1) if view_mode == "MTD" else pd.Timestamp(
        year=selected_date.year, month=1, day=1
    )
    data_period = actual_slice(start_date, selected_date)
    rows = []
    for code in process_codes:
        subset = data_period[data_period["공정코드"] == code]
        actual_sum = subset[actual_col].sum()
        good_sum = subset[good_col].sum()
        prod_sum = subset["생산수량"].sum()
        bad_sum = subset["불량수량"].sum()
        sample_sum = subset["샘플수량"].sum()
        workdays_actual = subset[subset["실제근무"]].shape[0]
        target_daily = target_lookup.get(code, 0)
        target_ref = target_daily * workdays_value
        target_actual = target_daily * workdays_actual
        rate_ref = actual_sum / target_ref if target_ref else pd.NA
        rate_act = actual_sum / target_actual if target_actual else pd.NA
        total_out = good_sum + bad_sum + sample_sum
        denom = max(prod_sum, total_out) if total_out else prod_sum
        yield_raw = actual_sum / prod_sum if prod_sum else pd.NA
        yield_adj = actual_sum / denom if denom else pd.NA
        anomaly_count = (
            subset["생산수량"]
            < (subset[good_col] + subset["불량수량"] + subset["샘플수량"])
        ).sum()
        if target_mode == "기준자료 근무일수 목표":
            target = target_ref
            rate = rate_ref
        elif target_mode == "실제근무일 목표":
            target = target_actual
            rate = rate_act
        else:
            target = target_ref
            rate = rate_ref
        rows.append(
            {
                "공정": process_display(code),
                "기준자료근무일수": workdays_value,
                "실제근무일수": workdays_actual,
                "일일_생산목표량": target_daily,
                "목표": target,
                "월목표_기준자료": target_ref,
                "월목표_실제근무": target_actual,
                "실적": actual_sum,
                "달성율": rate,
                "달성율_기준자료": rate_ref,
                "달성율_실제근무": rate_act,
                "생산수량": prod_sum,
                "수율(원칙)": yield_raw,
                "수율(보정)": yield_adj,
                "불량수량": bad_sum,
                "샘플수량": sample_sum,
                "수율이상건수": int(anomaly_count),
                "수율이상여부": "Y" if anomaly_count else "",
                "이상원인": "분모불일치" if anomaly_count else "",
            }
        )
    return pd.DataFrame(rows)


def build_excel_report(title, report_date, created_at, summary_df, table_df):
    from io import BytesIO

    output = BytesIO()
    try:
        import openpyxl  # noqa: F401
    except Exception:
        return None, "openpyxl not available"

    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Report", index=False, startrow=4)
            start_row = 6 + len(summary_df)
            table_df.to_excel(writer, sheet_name="Report", index=False, startrow=start_row)
            ws = writer.sheets["Report"]
            ws["A1"] = title
            ws["A2"] = f"보고 기준일: {report_date}"
            ws["A3"] = f"작성일시: {created_at}"
    except Exception as exc:  # noqa: BLE001
        return None, str(exc)
    output.seek(0)
    return output, None

def compute_monthly_kpis(
    actuals_df,
    targets_df,
    workdays_df,
    selected_year,
    selected_month,
    process_codes,
    view_mode,
    actual_col="실적",
):
    months = list(range(1, selected_month + 1)) if view_mode == "연간 누적(YTD)" else [selected_month]
    actual_col = actual_col if actual_col in actuals_df.columns else "실적"
    actuals = actuals_df[
        (actuals_df["연"] == selected_year)
        & (actuals_df["월"].isin(months))
        & (actuals_df["공정코드"].isin(process_codes))
    ]
    targets = targets_df[
        (targets_df["연"] == selected_year)
        & (targets_df["월"].isin(months))
        & (targets_df["공정코드"].isin(process_codes))
    ]
    workdays = workdays_df[
        (workdays_df["연"] == selected_year) & (workdays_df["월"].isin(months))
    ]

    base = targets.merge(workdays, on=["연", "월"], how="left")
    base = base.merge(
        actuals[
            [
                "연",
                "월",
                "공정코드",
                actual_col,
                "실적_양품",
                "생산수량",
                "불량수량",
                "샘플수량",
                "실제근무일수",
            ]
        ],
        on=["연", "월", "공정코드"],
        how="left",
    )
    base["실적"] = base[actual_col].fillna(0)
    base["실적_양품"] = base["실적_양품"].fillna(0)
    base["생산수량"] = base["생산수량"].fillna(0)
    base["불량수량"] = base["불량수량"].fillna(0)
    base["샘플수량"] = base["샘플수량"].fillna(0)
    base["실제근무일수"] = base["실제근무일수"].fillna(0)
    base["기준자료근무일수"] = base["기준자료근무일수"].fillna(0)
    base["기준근무일수"] = base["기준근무일수"].fillna(0)
    base["일일_생산목표량"] = base["일일_생산목표량"].fillna(0)

    base["월목표_기준자료"] = base["일일_생산목표량"] * base["기준자료근무일수"]
    base["월목표_실제근무"] = base["일일_생산목표량"] * base["실제근무일수"]

    if view_mode == "연간 누적(YTD)":
        grouped = (
            base.groupby("공정코드", dropna=False)
            .agg(
                기준자료근무일수=("기준자료근무일수", "sum"),
                실제근무일수=("실제근무일수", "sum"),
                일일_생산목표량=("일일_생산목표량", "mean"),
                월목표_기준자료=("월목표_기준자료", "sum"),
                월목표_실제근무=("월목표_실제근무", "sum"),
                실적=("실적", "sum"),
                실적_양품=("실적_양품", "sum"),
                생산수량=("생산수량", "sum"),
                불량수량=("불량수량", "sum"),
                샘플수량=("샘플수량", "sum"),
            )
            .reset_index()
        )
    else:
        grouped = (
            base.groupby("공정코드", dropna=False)
            .agg(
                기준자료근무일수=("기준자료근무일수", "max"),
                실제근무일수=("실제근무일수", "max"),
                일일_생산목표량=("일일_생산목표량", "max"),
                월목표_기준자료=("월목표_기준자료", "sum"),
                월목표_실제근무=("월목표_실제근무", "sum"),
                실적=("실적", "sum"),
                실적_양품=("실적_양품", "sum"),
                생산수량=("생산수량", "sum"),
                불량수량=("불량수량", "sum"),
                샘플수량=("샘플수량", "sum"),
            )
            .reset_index()
        )

    grouped["총출력"] = (
        grouped["실적_양품"] + grouped["불량수량"] + grouped["샘플수량"]
    )
    grouped["수율_원칙"] = grouped["실적"] / grouped["생산수량"].replace(0, pd.NA)
    denom = grouped[["생산수량", "총출력"]].max(axis=1).replace(0, pd.NA)
    grouped["수율_보정"] = grouped["실적"] / denom
    grouped["수율이상여부"] = grouped["생산수량"] < grouped["총출력"]
    grouped["수율"] = grouped["수율_원칙"]
    grouped["달성율_기준자료"] = grouped["실적"] / grouped["월목표_기준자료"].replace(0, pd.NA)
    grouped["달성율_실제근무"] = grouped["실적"] / grouped["월목표_실제근무"].replace(0, pd.NA)
    grouped["공정"] = grouped["공정코드"].apply(process_display)
    return grouped, base


def compute_yoy_metrics(grouped_2026, grouped_2025):
    merged = grouped_2026.merge(
        grouped_2025,
        on="공정코드",
        how="left",
        suffixes=("_2026", "_2025"),
    )
    merged["YoY_실적증감"] = merged["실적_2026"] - merged["실적_2025"].fillna(0)
    merged["YoY_실적증감%"] = merged["실적_2026"] / merged["실적_2025"].replace(0, pd.NA) - 1
    merged["YoY_수율증감"] = merged["수율_2026"] - merged["수율_2025"]
    merged["YoY_달성율증감_기준자료"] = (
        merged["달성율_기준자료_2026"] - merged["달성율_기준자료_2025"]
    )
    merged["YoY_달성율증감_실제근무"] = (
        merged["달성율_실제근무_2026"] - merged["달성율_실제근무_2025"]
    )
    return merged


def main():
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
    except Exception:  # noqa: BLE001
        get_script_run_ctx = None

    if get_script_run_ctx and get_script_run_ctx() is None:
        print("이 파일은 Streamlit으로 실행해야 합니다.")
        print("streamlit run streamlit_dashboard.py")
        return

    st.set_page_config(page_title="S관(3공장) 계획 대시보드", layout="wide")
    st.title("S관(3공장) 계획 대시보드")

    data_path = Path(DATA_FILE)
    if not data_path.exists():
        st.error(f"파일을 찾을 수 없습니다: {DATA_FILE}")
        return
    try:
        updated_at = datetime.fromtimestamp(data_path.stat().st_mtime)
        st.caption(f"업데이트 일자: {updated_at:%Y-%m-%d}")
    except OSError:
        pass

    try:
        df = load_data(data_path)
    except Exception as exc:  # noqa: BLE001
        st.error(f"데이터를 불러오지 못했습니다: {exc}")
        return

    fx_path = Path(FX_FILE)
    fx_map = {}
    if fx_path.exists():
        try:
            fx_df = load_fx(fx_path)
            fx_df.columns = [c.strip() for c in fx_df.columns]
            fx_df["기준"] = fx_df["기준"].astype(str).str.strip()
            if "원" not in fx_df.columns and "\\" in fx_df.columns:
                fx_df = fx_df.rename(columns={"\\": "원"})
            fx_map = fx_df.set_index("기준")[["원", "$"]].to_dict(orient="index")
        except Exception as exc:  # noqa: BLE001
            st.warning(f"환율 파일을 불러오지 못했습니다: {exc}")
    else:
        st.warning(f"환율 파일을 찾을 수 없습니다: {FX_FILE}")

    rename_map = {
        "잔여수주수량": "잔여수주량",
        "누수규격검사": "누수규격",
        "제품재고": "완제품",
        "하이드레이션/전면검사필요량": "수화필요량",
        "접착/멸균필요량": "접착필요량",
        "누수/규격검사필요량": "누수/규격필요량",
    }
    df = df.rename(columns=rename_map)

    required_cols = {"신규분류코드", "출고예상일", "생산필요량"}
    missing = required_cols - set(df.columns)
    if missing:
        st.error(f"필수 컬럼이 없습니다: {', '.join(sorted(missing))}")
        return

    df["출고예상일"] = pd.to_datetime(df["출고예상일"], errors="coerce")
    df["생산필요량"] = pd.to_numeric(df["생산필요량"], errors="coerce").fillna(0)
    if "수량" in df.columns:
        df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0)

    st.markdown(
        """
        <style>
        .block-separator { border-top: 1px solid #D9D9D9; margin: 10px 0; }
        [data-testid="stDataFrame"] th { text-align: center !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    category_values = sorted(
        value for value in df["신규분류코드"].dropna().unique().tolist() if value != ""
    )

    with st.sidebar:
        st.header("필터")
        st.checkbox(
            "빠른 로딩(서식 최소화)",
            value=True,
            key=FAST_MODE_KEY,
            help="표 서식/정렬을 줄이고 로딩 속도를 높입니다.",
        )
        selected_categories = st.multiselect(
            "신규분류코드",
            options=category_values,
            default=category_values,
        )
        search_keyword = st.text_input(
            "전체 검색 (품명/이니셜/품목코드)",
            value="",
            placeholder="검색어 입력",
        )
        date_range = None
        if "출고예상일" in df.columns:
            date_min = df["출고예상일"].dropna().min()
            date_max = df["출고예상일"].dropna().max()
            if pd.notna(date_min) and pd.notna(date_max):
                date_range = st.date_input(
                    "출고예상일 범위",
                    value=(date_min.date(), date_max.date()),
                )
        ship_scope = None
        if "수출국가" in df.columns:
            ship_scope = st.selectbox("수출/국내", ["전체", "수출", "국내", "예외"], index=0)
        view_mode = st.radio("보기", ["생산필요만", "전체"], horizontal=True)
        sort_by_ship = True

    filtered = df.copy()
    if selected_categories:
        filtered = filtered[filtered["신규분류코드"].isin(selected_categories)]
    else:
        filtered = filtered.iloc[0:0]
    if ship_scope and ship_scope != "전체" and "수출국가" in filtered.columns:
        scope_series = filtered["수출국가"].astype(str).str.strip()
        scope_series = scope_series.apply(
            lambda value: "국내"
            if value == "국내"
            else ("예외" if value.startswith("예외:") or value == "예외" else "수출")
        )
        filtered = filtered[scope_series == ship_scope]
    if date_range and isinstance(date_range, tuple) and len(date_range) == 2:
        start_date = pd.to_datetime(date_range[0])
        end_date = pd.to_datetime(date_range[1])
        filtered = filtered[
            (filtered["출고예상일"].notna())
            & (filtered["출고예상일"] >= start_date)
            & (filtered["출고예상일"] <= end_date)
        ]
    if search_keyword.strip():
        search_cols = [
            col
            for col in ["품명", "이니셜", "품목코드", "판매코드", "수주번호"]
            if col in filtered.columns
        ]
        if search_cols:
            mask = (
                filtered[search_cols]
                .astype(str)
                .apply(
                    lambda row: row.str.contains(search_keyword, case=False, na=False).any(),
                    axis=1,
                )
            )
            filtered = filtered[mask]
    filtered_base = filtered.copy()

    if view_mode == "생산필요만":
        filtered = filtered[filtered["생산필요량"] > 0]

    if sort_by_ship and "출고예상일" in filtered.columns:
        filtered = filtered.sort_values(by="출고예상일", ascending=True)

    if DEBUG_DUPLICATES:
        debug_duplicate_orders(filtered_base, "filtered_base")

    (
        tab_all,
        tab_production,
        tab_c_support,
        tab_inventory,
        tab_goal,
    ) = st.tabs(
        ["S관 종합현황", "생산계획", "C관접착지원", "재고현황", "S관 공장목표현황"]
    )

    with tab_all:
        summary_df = filtered_base

        def add_fx_amounts(frame):
            result = frame.copy()
            result["수주금액"] = pd.to_numeric(
                result.get("수주금액", 0).astype(str).str.replace(",", ""),
                errors="coerce",
            ).fillna(0.0)
            result["원환산"] = 0.0
            result["달러환산"] = 0.0
            if "화폐" in result.columns and fx_map:
                for currency, fx in fx_map.items():
                    mask = result["화폐"].astype(str).str.strip() == currency
                    if not mask.any():
                        continue
                    result.loc[mask, "원환산"] = result.loc[mask, "수주금액"] * float(
                        fx.get("원", 0)
                    )
                    result.loc[mask, "달러환산"] = result.loc[mask, "수주금액"] * float(
                        fx.get("$", 0)
                    )
                krw_mask = result["화폐"].astype(str).str.strip() == "KRW"
                if krw_mask.any():
                    result.loc[krw_mask, "원환산"] = result.loc[krw_mask, "수주금액"]
                    result.loc[krw_mask, "달러환산"] = result.loc[krw_mask, "수주금액"] / 1300
            return result

        month_source = "출고예상일" if "출고예상일" in summary_df.columns else None
        if month_source:
            month_df = summary_df.copy()
            month_df[month_source] = pd.to_datetime(
                month_df[month_source], errors="coerce"
            )
            month_df = month_df[month_df[month_source].notna()]
            month_df["월"] = month_df[month_source].dt.to_period("M").astype(str)
            if "수출국가" in month_df.columns:
                month_df["수출/국내"] = month_df["수출국가"].astype(str).str.strip()
                def map_scope(value):
                    if value == "국내":
                        return "국내"
                    if value.startswith("예외:"):
                        mapped = value.replace("예외:", "", 1).strip()
                        return mapped if mapped else "예외"
                    if value == "예외":
                        return "예외"
                    return "수출"
                month_df["수출/국내"] = month_df["수출/국내"].apply(map_scope)
            else:
                month_df["수출/국내"] = "수출"
            month_df["수량"] = pd.to_numeric(
                month_df.get("수량", 0), errors="coerce"
            ).fillna(0)
            month_df["생산필요량"] = pd.to_numeric(
                month_df.get("생산필요량", 0), errors="coerce"
            ).fillna(0)

            fx_month = add_fx_amounts(month_df)

            total_counts = (
                month_df.groupby(["월", "수출/국내"])
                .agg(
                    이니셜수=("이니셜", "nunique"),
                    품목수=("품명", "nunique" if "품명" in month_df.columns else "count"),
                    규격수=("판매코드", "nunique"),
                )
                .reset_index()
            )
            remaining_df = month_df[month_df["생산필요량"] > 0]
            remain_counts = (
                remaining_df.groupby(["월", "수출/국내"])
                .agg(
                    잔여이니셜수=("이니셜", "nunique"),
                    잔여품목수=("품명", "nunique" if "품명" in remaining_df.columns else "count"),
                    잔여규격수=("판매코드", "nunique"),
                )
                .reset_index()
            )

            monthly = (
                month_df.groupby(["월", "수출/국내"], dropna=False)[["수량", "생산필요량"]]
                .sum()
                .reset_index()
            )
            if "이니셜" in fx_month.columns:
                amounts = (
                    fx_month.groupby(
                        ["월", "수출/국내", "이니셜"], dropna=False
                    )[["원환산", "달러환산"]]
                    .max()
                    .reset_index()
                    .groupby(["월", "수출/국내"], dropna=False)[["원환산", "달러환산"]]
                    .sum()
                    .reset_index()
                )
            else:
                amounts = (
                    fx_month.groupby(["월", "수출/국내"], dropna=False)[
                        ["원환산", "달러환산"]
                    ]
                    .sum()
                    .reset_index()
                )
            monthly = monthly.merge(amounts, on=["월", "수출/국내"], how="left")
            monthly = monthly.merge(total_counts, on=["월", "수출/국내"], how="left")
            monthly = monthly.merge(remain_counts, on=["월", "수출/국내"], how="left")

            monthly["진도율"] = (
                (1 - (monthly["생산필요량"] / monthly["수량"])) * 100
            ).fillna(0)

            def fmt_int(value):
                try:
                    return f"{int(round(float(value))):,}"
                except (TypeError, ValueError):
                    return "0"

            def fmt_ratio(value):
                try:
                    return f"{float(value):.2f}"
                except (TypeError, ValueError):
                    return "0.00"

            monthly["이니셜수"] = (
                monthly["이니셜수"].fillna(0).apply(fmt_int)
                + "("
                + monthly["잔여이니셜수"].fillna(0).apply(fmt_int)
                + ")"
            )
            monthly["품목수"] = (
                monthly["품목수"].fillna(0).apply(fmt_int)
                + "("
                + monthly["잔여품목수"].fillna(0).apply(fmt_int)
                + ")"
            )
            monthly["규격수"] = (
                monthly["규격수"].fillna(0).apply(fmt_int)
                + "("
                + monthly["잔여규격수"].fillna(0).apply(fmt_int)
                + ")"
            )

            monthly = monthly.rename(
                columns={
                    "수량": "월 수주수량",
                    "원환산": "수주금액(원)",
                    "달러환산": "수주금액($)",
                }
            )
            monthly = monthly.sort_values(["월", "수출/국내"])
            total_qty = monthly["월 수주수량"].sum()
            total_need = monthly["생산필요량"].sum()
            total_ratio = (1 - (total_need / total_qty)) * 100 if total_qty else 0
            total_krw = monthly["수주금액(원)"].sum()
            total_usd = monthly["수주금액($)"].sum()

            monthly["월 수주수량"] = monthly["월 수주수량"].apply(fmt_int)
            monthly["생산필요량"] = monthly["생산필요량"].apply(fmt_int)
            monthly["진도율"] = monthly["진도율"].apply(lambda v: f"{fmt_ratio(v)}%")
            monthly["수주금액(원)"] = monthly["수주금액(원)"].apply(fmt_int)
            monthly["수주금액($)"] = monthly["수주금액($)"].apply(fmt_int)
            monthly = monthly.fillna("")
            monthly = monthly[monthly["월 수주수량"] != "0"]
            monthly = monthly[
                [
                    "수출/국내",
                    "월",
                    "월 수주수량",
                    "생산필요량",
                    "진도율",
                    "이니셜수",
                    "품목수",
                    "규격수",
                    "수주금액(원)",
                    "수주금액($)",
                ]
            ]
            st.subheader("S관 수주접수 현황")
            kpi = st.columns(5)
            kpi[0].metric("월 수주수량 합계", fmt_int(total_qty))
            kpi[1].metric("생산필요량 합계", fmt_int(total_need))
            kpi[2].metric("진도율", f"{fmt_ratio(total_ratio)}%")
            kpi[3].metric("수주금액(원)", fmt_int(total_krw))
            kpi[4].metric("수주금액($)", fmt_int(total_usd))
            def highlight_need_columns(data):
                styles = pd.DataFrame("", index=data.index, columns=data.columns)
                for col in data.columns:
                    if "필요량" in str(col):
                        styles.loc[:, col] = "color: #D0021B; font-weight: 600;"
                return styles

            monthly_styled = apply_alignment(
                monthly.style.apply(highlight_need_columns, axis=None),
                monthly,
            )
            st.dataframe(
                monthly_styled,
                width="stretch",
                height=calc_table_height(len(monthly), max_height=360, min_height=180),
            )

        st.subheader("상세 목록")
        hide_columns = {
            "잔여수량",
            "사출창고",
            "분리창고",
            "검사접착",
            "누수규격",
            "사출필요량",
            "분리필요량",
            "수화필요량",
            "접착필요량",
            "누수/규격필요량",
        }
        detail_cols = [col for col in filtered.columns if col not in hide_columns]
        search_text = st.text_input("상세 목록 검색", value="", placeholder="검색어 입력")
        detail_df = filtered[detail_cols].copy()
        if "수량" in detail_df.columns and "잔여수량" in detail_df.columns:
            cols = detail_df.columns.tolist()
            cols.remove("수량")
            if "잔여수량" in cols:
                idx = cols.index("잔여수량")
                cols.insert(idx, "수량")
            else:
                cols.append("수량")
            detail_df = detail_df[cols]
        if "출고예상일" in detail_df.columns:
            sort_dates = pd.to_datetime(detail_df["출고예상일"], errors="coerce")
            detail_df = (
                detail_df.assign(_sort_date=sort_dates)
                .sort_values("_sort_date", ascending=True)
                .drop(columns=["_sort_date"])
            )
        overdue_mask = None
        soon_mask = None
        if "출고예상일" in detail_df.columns:
            detail_df, overdue_mask, soon_mask = add_due_warning(detail_df, "출고예상일", 7)
        date_keywords = ("일자", "날짜", "출고", "납기")
        date_cols = [
            col
            for col in detail_df.columns
            if col != "출고예상일"
            and (
                any(key in col for key in date_keywords)
                or col.lower().endswith("date")
            )
        ]
        for col in date_cols:
            detail_df[col] = format_date_series(detail_df[col])
        numeric_cols = [
            col
            for col in detail_df.columns
            if any(key in col for key in ("수량", "단가", "금액", "필요량", "재고"))
        ]
        for col in numeric_cols:
            detail_df[col] = pd.to_numeric(
                detail_df[col].astype(str).str.replace(",", ""), errors="coerce"
            )
        detail_df = fill_object_na(detail_df)
        if search_text.strip():
            mask = detail_df.astype(str).apply(
                lambda row: row.str.contains(search_text, case=False, na=False).any(),
                axis=1,
            )
            detail_df = detail_df[mask]
        if overdue_mask is not None:
            overdue_mask = overdue_mask.reindex(detail_df.index, fill_value=False)
        if soon_mask is not None:
            soon_mask = soon_mask.reindex(detail_df.index, fill_value=False)

        def highlight_warning(data):
            styles = pd.DataFrame("", index=data.index, columns=data.columns)
            if "출고예상일" not in styles.columns:
                return styles
            for col in styles.columns:
                if "필요량" in str(col):
                    styles.loc[:, col] = "color: #D0021B; font-weight: 600;"
            for idx in data.index:
                if overdue_mask is not None and overdue_mask.loc[idx]:
                    styles.loc[idx, "출고예상일"] = "color: #D0021B; font-weight: 700;"
                elif soon_mask is not None and soon_mask.loc[idx]:
                    styles.loc[idx, "출고예상일"] = "color: #B7791F; font-weight: 700;"
            return styles

        format_dict = {}
        for col in detail_df.columns:
            if not pd.api.types.is_numeric_dtype(detail_df[col]):
                continue
            if "단가" in col or "율" in col or "비율" in col:
                format_dict[col] = "{:,.2f}"
            else:
                format_dict[col] = "{:,.0f}"
        styled = apply_alignment(
            detail_df.style.format(format_dict, na_rep="").apply(
                highlight_warning, axis=None
            ),
            detail_df,
        )
        selection = render_dataframe(
            styled,
            detail_df,
            width="stretch",
            height=calc_table_height(len(detail_df)),
            on_select="rerun",
            selection_mode="multi-row",
            key="detail_table",
        )
        if selection is not None:
            render_selection_sum(detail_df, selection.selection.rows, "선택 합계")

    def render_production_table(view_df, category_name, key_suffix, show_warning=False):
        st.markdown(f"**{category_name} 생산계획**")
        if view_df.empty:
            return

        aggregated = aggregate_production(view_df.copy())
        add_spec_columns(aggregated)
        if "출고예상일" in aggregated.columns:
            aggregated["출고예상일"] = pd.to_datetime(
                aggregated["출고예상일"], errors="coerce"
            )
            aggregated["_sort_date"] = aggregated["출고예상일"]
        aggregated = filter_need_rows(aggregated)
        if aggregated.empty:
            return
        overdue_mask = None
        soon_mask = None
        if show_warning:
            aggregated, overdue_mask, soon_mask = add_due_warning(
                aggregated, "출고예상일", 7
            )

        show_codes = st.checkbox(
            "코드 컬럼 펼치기 (품목코드, Q코드, R코드)",
            value=False,
            key=f"show_codes_prod_{key_suffix}",
        )

        stock_cols = [
            "사출창고",
            "분리창고",
            "검사접착",
            "누수규격",
            "완제품",
            "불용재고",
        ]
        need_cols = ["사출필요량", "분리필요량", "수화필요량", "접착필요량", "누수/규격필요량"]
        columns = ["품명", "생산품명", "신규분류코드"]
        if show_codes:
            columns.extend(["품목코드", "Q코드", "R코드"])
        columns.extend(
            ["파워", "실린더(ADD)", "축", "출고예상일", "수량", "잔여수주량", "생산필요량"]
        )
        columns.extend([col for col in stock_cols if col in aggregated.columns])
        columns.extend([col for col in need_cols if col in aggregated.columns])
        available = [col for col in columns if col in aggregated.columns]
        extra_cols = ["_sort_date"] if "_sort_date" in aggregated.columns else []
        view = aggregated[available + extra_cols].copy()

        if sort_by_ship and "_sort_date" in view.columns:
            view = view.sort_values(by="_sort_date", ascending=True)
        if "_sort_date" in view.columns:
            view = view.drop(columns=["_sort_date"])
        if "출고예상일" in view.columns:
            if show_warning:
                view["출고예상일"] = view["출고예상일"].fillna("")
            else:
                view["출고예상일"] = format_date_series(view["출고예상일"])

        search_text = st.text_input(
            "검색",
            value="",
            placeholder="검색어 입력",
            key=f"search_prod_{key_suffix}",
        )
        if search_text.strip():
            mask = view.astype(str).apply(
                lambda row: row.str.contains(search_text, case=False, na=False).any(),
                axis=1,
            )
            view = view[mask]

        view = fill_object_na(view)
        top_labels = []
        for col in view.columns:
            if col in stock_cols:
                top_labels.append("재고현황")
            elif col in need_cols:
                top_labels.append("생산필요량")
            else:
                top_labels.append("")
        view.columns = pd.MultiIndex.from_arrays([top_labels, view.columns])

        def border_styles(data):
            styles = pd.DataFrame("", index=data.index, columns=data.columns)

            def append_style(series, addition):
                return series.apply(lambda value: f"{value}; {addition}" if value else addition)

            stock_first = ("재고현황", stock_cols[0]) if stock_cols else None
            stock_last = ("재고현황", stock_cols[-1]) if stock_cols else None
            need_first = ("생산필요량", need_cols[0]) if need_cols else None
            need_last = ("생산필요량", need_cols[-1]) if need_cols else None
            if stock_first and stock_first in styles.columns:
                styles.loc[:, stock_first] = append_style(
                    styles.loc[:, stock_first], "border-left: 2px solid #777"
                )
            if stock_last and stock_last in styles.columns:
                styles.loc[:, stock_last] = append_style(
                    styles.loc[:, stock_last], "border-right: 2px solid #777"
                )
            if need_first and need_first in styles.columns:
                styles.loc[:, need_first] = append_style(
                    styles.loc[:, need_first], "border-left: 2px solid #777"
                )
            if need_last and need_last in styles.columns:
                styles.loc[:, need_last] = append_style(
                    styles.loc[:, need_last], "border-right: 2px solid #777"
                )
            for col in stock_cols:
                key = ("재고현황", col)
                if key in styles.columns:
                    styles.loc[:, key] = append_style(
                        styles.loc[:, key], "background-color: #EAF2FF"
                    )
            for col in need_cols:
                key = ("생산필요량", col)
                if key in styles.columns:
                    styles.loc[:, key] = append_style(
                        styles.loc[:, key], "background-color: #FFF1E6"
                    )
            for col in styles.columns:
                col_name = col[1] if isinstance(col, tuple) else col
                if "필요량" in str(col_name):
                    styles.loc[:, col] = append_style(
                        styles.loc[:, col], "color: #D0021B; font-weight: 600;"
                    )
            return styles

        header_styles = [
            {"selector": "th.col_heading.level0", "props": [("text-align", "center")]},
            {"selector": "th.col_heading.level1", "props": [("text-align", "center")]},
        ]
        sticky_idx = None
        for idx, (group, col) in enumerate(view.columns):
            if group == "재고현황":
                header_styles.append(
                    {
                        "selector": f"th.col_heading.level0.col{idx}",
                        "props": [("background-color", "#AFC8FF"), ("color", "#0B3B8C")],
                    }
                )
            elif group == "생산필요량":
                header_styles.append(
                    {
                        "selector": f"th.col_heading.level0.col{idx}",
                        "props": [("background-color", "#FFBE8C"), ("color", "#8A3B00")],
                    }
                )
            if col in stock_cols:
                header_styles.append(
                    {
                        "selector": f"th.col_heading.level1.col{idx}",
                        "props": [("background-color", "#C7DCFF"), ("color", "#0B3B8C")],
                    }
                )
            elif col in need_cols:
                header_styles.append(
                    {
                        "selector": f"th.col_heading.level1.col{idx}",
                        "props": [("background-color", "#FFD0B0"), ("color", "#8A3B00")],
                    }
                )
            if col == "품명" and sticky_idx is None:
                sticky_idx = idx

        if sticky_idx is not None:
            header_styles.append(
                {
                    "selector": f"th.col_heading.level0.col{sticky_idx}",
                    "props": [
                        ("position", "sticky"),
                        ("left", "0px"),
                        ("z-index", "5"),
                        ("background-color", "#F5F5F5"),
                    ],
                }
            )
            header_styles.append(
                {
                    "selector": f"th.col_heading.level1.col{sticky_idx}",
                    "props": [
                        ("position", "sticky"),
                        ("left", "0px"),
                        ("z-index", "6"),
                        ("background-color", "#F5F5F5"),
                    ],
                }
            )
            header_styles.append(
                {
                    "selector": f"td.col{sticky_idx}",
                    "props": [
                        ("position", "sticky"),
                        ("left", "0px"),
                        ("z-index", "4"),
                        ("background-color", "#FFFFFF"),
                    ],
                }
            )

        if overdue_mask is not None:
            overdue_mask = overdue_mask.reindex(aggregated.index, fill_value=False)
            overdue_mask = overdue_mask.loc[view.index]
        if soon_mask is not None:
            soon_mask = soon_mask.reindex(aggregated.index, fill_value=False)
            soon_mask = soon_mask.loc[view.index]

        def highlight_warning(data):
            styles = pd.DataFrame("", index=data.index, columns=data.columns)
            if overdue_mask is None and soon_mask is None:
                return styles
            due_key = ("", "출고예상일") if ("", "출고예상일") in styles.columns else None
            if due_key is None:
                return styles
            for idx in data.index:
                if overdue_mask is not None and overdue_mask.loc[idx]:
                    styles.loc[idx, due_key] = "color: #D0021B; font-weight: 700;"
                elif soon_mask is not None and soon_mask.loc[idx]:
                    styles.loc[idx, due_key] = "color: #B7791F; font-weight: 700;"
            return styles

        format_dict = {}
        for col in view.columns:
            series = view[col]
            if not pd.api.types.is_numeric_dtype(series):
                continue
            name = col[1] if isinstance(col, tuple) else col
            if "단가" in name or "율" in name or "비율" in name:
                format_dict[col] = "{:,.2f}"
            else:
                format_dict[col] = "{:,.0f}"
        styled = apply_alignment(
            view.style.format(format_dict, na_rep="")
            .apply(border_styles, axis=None)
            .apply(highlight_warning, axis=None)
            .set_table_styles(header_styles),
            view,
        )
        selection = render_dataframe(
            styled,
            view,
            width="stretch",
            height=calc_table_height(len(view)),
            on_select="rerun",
            selection_mode="multi-row",
            key=f"prod_table_{key_suffix}",
        )
        if selection is not None:
            render_selection_sum(view, selection.selection.rows, "선택 합계")

    def render_injection_summary(view_df, title):
        st.markdown(f"**{title}**")
        if "R코드" not in view_df.columns:
            st.info("R코드 컬럼이 없습니다.")
            return
        summary = view_df.copy()
        summary = summary[summary["R코드"].astype(str).str.strip() != ""]
        if "품명" not in summary.columns:
            summary["품명"] = ""
        if "사출창고" not in summary.columns:
            summary["사출창고"] = 0
        for col in ("생산필요량", "사출필요량", "사출창고"):
            if col in summary.columns:
                summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0)
        if "사출필요량" in summary.columns:
            summary = summary[summary["사출필요량"] > 0]
        if summary.empty:
            st.info("사출필요량이 있는 품목이 없습니다.")
            return

        stock_by_r = summary.drop_duplicates(subset=["R코드"])[["R코드", "사출창고"]]
        stock_by_r["사출창고"] = pd.to_numeric(
            stock_by_r["사출창고"], errors="coerce"
        ).fillna(0)

        summary["출고예상일"] = pd.to_datetime(summary["출고예상일"], errors="coerce")
        grouped = (
            summary.groupby("R코드", dropna=True)
            .agg(
                {
                    "품명": representative_name,
                    "생산필요량": "sum",
                    "사출필요량": "sum",
                    "출고예상일": "min",
                }
            )
            .reset_index()
        )
        grouped = grouped.merge(stock_by_r, on="R코드", how="left")

        specs = grouped["R코드"].apply(parse_spec_from_code).apply(pd.Series)
        specs.columns = ["파워", "실린더(ADD)", "축"]
        grouped = pd.concat([grouped, specs], axis=1)
        grouped["_sort_date"] = pd.to_datetime(grouped["출고예상일"], errors="coerce")
        grouped, overdue_mask, soon_mask = add_due_warning(grouped, "출고예상일", 7)
        ordered_cols = [
            "R코드",
            "품명",
            "파워",
            "실린더(ADD)",
            "축",
            "생산필요량",
            "사출창고",
            "사출필요량",
            "출고예상일",
        ]
        keep_cols = [col for col in ordered_cols if col in grouped.columns]
        if "_sort_date" in grouped.columns:
            keep_cols.append("_sort_date")
        grouped = grouped[keep_cols]
        if "_sort_date" in grouped.columns:
            grouped = grouped.sort_values(by="_sort_date", ascending=True)
            grouped = grouped.drop(columns=["_sort_date"])
        if overdue_mask is not None:
            overdue_mask = overdue_mask.reindex(grouped.index, fill_value=False)
        if soon_mask is not None:
            soon_mask = soon_mask.reindex(grouped.index, fill_value=False)

        def highlight_injection(data):
            styles = pd.DataFrame("", index=data.index, columns=data.columns)
            if "사출창고" in styles.columns:
                styles.loc[:, "사출창고"] = "background-color: #EAF2FF"
            for col in ("생산필요량", "사출필요량"):
                if col in styles.columns:
                    styles.loc[:, col] = "background-color: #FFF1E6; color: #D0021B; font-weight: 600;"
            for col in styles.columns:
                if "필요량" in str(col):
                    styles.loc[:, col] = "color: #D0021B; font-weight: 600;"
            if "출고예상일" in styles.columns:
                for idx in data.index:
                    if overdue_mask is not None and overdue_mask.loc[idx]:
                        styles.loc[idx, "출고예상일"] = "color: #D0021B; font-weight: 700;"
                    elif soon_mask is not None and soon_mask.loc[idx]:
                        styles.loc[idx, "출고예상일"] = "color: #B7791F; font-weight: 700;"
            return styles

        format_dict = {}
        for col in grouped.columns:
            if not pd.api.types.is_numeric_dtype(grouped[col]):
                continue
            if "단가" in col or "율" in col or "비율" in col:
                format_dict[col] = "{:,.2f}"
            else:
                format_dict[col] = "{:,.0f}"
        styled = apply_alignment(
            grouped.style.format(format_dict, na_rep="").apply(
                highlight_injection, axis=None
            ),
            grouped,
        )
        selection = render_dataframe(
            styled,
            grouped,
            width="stretch",
            height=calc_table_height(len(grouped)),
            on_select="rerun",
            selection_mode="multi-row",
            key=f"injection_table_{title}",
        )
        if selection is not None:
            render_selection_sum(grouped, selection.selection.rows, "선택 합계")

    has_category = "신규분류코드" in filtered_base.columns
    color_mask = (
        filtered_base["신규분류코드"].astype(str).str.contains("color", case=False, na=False)
        if has_category
        else pd.Series([False] * len(filtered_base), index=filtered_base.index)
    )

    def render_category_tabs(view_df, label, key_prefix):
        view_df = filter_need_rows(view_df)
        if view_df.empty:
            return
        st.subheader(f"{label} 생산계획")
        if "신규분류코드" in view_df.columns:
            categories = []
            for value in view_df["신규분류코드"].dropna().unique().tolist():
                if not str(value).strip():
                    continue
                subset = filter_need_rows(view_df[view_df["신규분류코드"] == value])
                if not subset.empty:
                    categories.append(value)
            categories = sorted(categories)
        else:
            categories = []
        main_label = f"{label}현황"
        tab_names = [main_label, "사출코드"] + categories if categories else [main_label, "사출코드"]
        tabs = st.tabs(tab_names)
        with tabs[0]:
            render_production_table(
                view_df, main_label, f"{key_prefix}_all", show_warning=True
            )
        with tabs[1]:
            st.subheader("사출공정 (R코드 집계)")
            render_injection_summary(view_df, f"{label} 사출공정 요약")
        for idx, category in enumerate(categories, start=2):
            with tabs[idx]:
                render_production_table(
                    view_df[view_df["신규분류코드"] == category],
                    f"{category}",
                    f"{key_prefix}_{category}",
                    show_warning=True,
                )

    with tab_production:
        color_tab, clear_tab = st.tabs(["Color", "Clear"])
        with color_tab:
            render_category_tabs(filtered_base[color_mask], "Color", "color")
        with clear_tab:
            render_category_tabs(filtered_base[~color_mask], "Clear", "clear")

    with tab_c_support:
        # TEMP: C관접착지원 탭 (임시 운영)
        st.subheader("C관접착지원")
        support_path = Path("깃허브_C관접착지원_요약.csv")
        if not support_path.exists():
            st.warning("깃허브_C관접착지원_요약.csv 파일을 찾을 수 없습니다.")
        else:
            try:
                support_df = load_data(support_path)
            except Exception as exc:  # noqa: BLE001
                st.warning(f"깃허브_C관접착지원 파일을 불러오지 못했습니다: {exc}")
                support_df = None

        if support_df is None:
            st.info("C관접착지원 데이터를 표시할 수 없습니다.")
        else:
            support_df = support_df.rename(columns={col: col.strip() for col in support_df.columns})
            support_df = support_df.dropna(how="all")
            required_cols = {"제품명", "제품코드", "생산부족수량", "납기"}
            missing_cols = required_cols - set(support_df.columns)
            if missing_cols:
                st.warning(
                    f"깃허브_C관접착지원 필수 컬럼이 없습니다: {', '.join(sorted(missing_cols))}"
                )
                support_df = None

        if support_df is None:
            st.info("C관접착지원 데이터를 표시할 수 없습니다.")
        else:
            for col in ["생산부족수량", "생산실적", "진도율"]:
                if col in support_df.columns:
                    support_df[col] = pd.to_numeric(
                        support_df[col].astype(str).str.replace(",", "").str.strip(),
                        errors="coerce",
                    ).fillna(0)
            if "생산부족수량" in support_df.columns:
                support_df["생산부족수량"] = support_df["생산부족수량"].abs()

            summary_name_count = support_df["제품명"].nunique()
            summary_spec_count = support_df["제품코드"].nunique()
            total_shortage = support_df["생산부족수량"].sum()
            total_actual = (
                support_df["생산실적"].sum() if "생산실적" in support_df.columns else 0
            )
            total_ratio = (total_actual / total_shortage) if total_shortage else 0

            kpi = st.columns(5)
            kpi[0].metric("품명", f"{summary_name_count:,}")
            kpi[1].metric("규격수", f"{summary_spec_count:,}")
            kpi[2].metric("요청수량", f"{int(total_shortage):,}")
            kpi[3].metric("생산실적", f"{int(total_actual):,}")
            kpi[4].metric("진도율", f"{total_ratio * 100:.2f}%")

            display_cols = [
                "제품코드",
                "제품명",
                "파워",
                "생산부족수량",
                "납기",
                "생산실적",
                "진도율",
            ]
            available_cols = [col for col in display_cols if col in support_df.columns]
            view = support_df[available_cols].copy()
            view = fill_object_na(view)
            if "진도율" in view.columns:
                view["진도율"] = view["진도율"].apply(lambda v: f"{v * 100:.2f}%")
            format_dict = {}
            for col in view.columns:
                if pd.api.types.is_numeric_dtype(view[col]) and col != "진도율":
                    format_dict[col] = "{:,.0f}"
            styled = apply_alignment(
                view.style.format(format_dict, na_rep=""),
                view,
            )
            st.dataframe(
                styled,
                width="stretch",
                height=calc_table_height(len(view), max_height=520),
            )

    with tab_inventory:
        st.subheader("재고현황")
        inventory_path = Path("깃허브_재고현황_요약.csv")
        if not inventory_path.exists():
            st.warning("깃허브_재고현황_요약.csv 파일을 찾을 수 없습니다.")
        else:
            try:
                inventory_df = load_data(inventory_path)
            except Exception as exc:  # noqa: BLE001
                st.warning(f"깃허브_재고현황_요약.csv 파일을 불러오지 못했습니다: {exc}")
                inventory_df = None

            if inventory_df is None:
                st.info("재고현황 데이터를 표시할 수 없습니다.")
            else:
                inventory_df = inventory_df.rename(
                    columns={col: col.strip() for col in inventory_df.columns}
                )
                required = {"품목코드", "사출창고", "분리창고", "검사접착", "누수규격", "완제품"}
                missing = required - set(inventory_df.columns)
                if missing:
                    st.warning(
                        f"깃허브_재고현황_요약.csv 필수 컬럼이 없습니다: {', '.join(sorted(missing))}"
                    )
                    inventory_df = None

            if inventory_df is None:
                st.info("재고현황 데이터를 표시할 수 없습니다.")
            else:
                numeric_cols = [
                    "사출창고",
                    "분리창고",
                    "검사접착",
                    "누수규격",
                    "완제품",
                    "불용재고",
                ]
                for col in [name for name in numeric_cols if name in inventory_df.columns]:
                    inventory_df[col] = pd.to_numeric(
                        inventory_df[col].astype(str).str.replace(",", "").str.strip(),
                        errors="coerce",
                    ).fillna(0)
                name_col = "품명" if "품명" in inventory_df.columns else None
                pivoted = inventory_df.copy()

            if inventory_df is None:
                st.info("재고현황 데이터를 표시할 수 없습니다.")
            else:
                needs = None
                if "품목코드" in df.columns and "생산필요량" in df.columns:
                    needs = (
                        df.groupby("품목코드", dropna=False)["생산필요량"]
                        .sum()
                        .reset_index()
                    )
                    pivoted = pivoted.merge(needs, on="품목코드", how="left")
                else:
                    pivoted["생산필요량"] = 0
                pivoted["생산필요량"] = pd.to_numeric(
                    pivoted["생산필요량"], errors="coerce"
                ).fillna(0)
                pivoted["잔여수량"] = pivoted["완제품"] - pivoted["생산필요량"]
                pivoted.loc[pivoted["잔여수량"] <= 0, "잔여수량"] = pd.NA

                search_text = st.text_input(
                    "재고 검색",
                    value="",
                    placeholder="품목코드/품명 검색",
                    key="inventory_search",
                )
                if search_text.strip():
                    search_cols = ["품목코드"]
                    if name_col:
                        search_cols.append("품명")
                    mask = (
                        pivoted[search_cols]
                        .astype(str)
                        .apply(
                            lambda row: row.str.contains(search_text, case=False, na=False).any(),
                            axis=1,
                        )
                    )
                    pivoted = pivoted[mask]

                columns = [
                    "품목코드",
                    "품명",
                    "사출창고",
                    "분리창고",
                    "검사접착",
                    "누수규격",
                    "완제품",
                    "불용재고",
                    "생산필요량",
                    "잔여수량",
                ]
                available = [col for col in columns if col in pivoted.columns]
                view = pivoted[available].copy()
                view = fill_object_na(view)

                format_dict = {}
                for col in view.columns:
                    if pd.api.types.is_numeric_dtype(view[col]):
                        format_dict[col] = "{:,.0f}"
                styled = apply_alignment(
                    view.style.format(format_dict, na_rep=""),
                    view,
                )
                selection = render_dataframe(
                    styled,
                    view,
                    width="stretch",
                    height=calc_table_height(len(view), max_height=650),
                    on_select="rerun",
                    selection_mode="multi-row",
                    key="inventory_table",
                )
                if selection is not None:
                    render_selection_sum(view, selection.selection.rows, "선택 합계")



    with tab_goal:
        header_placeholder = st.empty()

        goal_summary = load_goal_summary()
        actuals_raw = load_actuals()
        targets_raw = load_targets()
        workdays_raw = load_workdays()

        targets = preprocess_targets(targets_raw)
        workdays = preprocess_workdays(workdays_raw)

        if not goal_summary.empty:
            daily_actuals = goal_summary.copy()
            actuals = aggregate_actuals_from_daily(daily_actuals)
        else:
            actuals = preprocess_actuals(actuals_raw)
            daily_actuals = preprocess_actuals_daily(actuals_raw)

        if actuals.empty or targets.empty or workdays.empty or daily_actuals.empty:
            st.warning("S관 공장목표현황 데이터를 불러올 수 없습니다.")
        else:
            latest_date = daily_actuals["일자"].max()
            latest_date = latest_date.date() if pd.notna(latest_date) else datetime.now().date()
            default_compare = latest_date - pd.Timedelta(days=1)

            with st.sidebar.expander("S관 공장목표현황 필터", expanded=True):
                selected_date = st.date_input(
                    "기준일",
                    value=latest_date,
                    key="goal_date",
                )
                compare_date = st.date_input(
                    "비교일(전일)",
                    value=default_compare,
                    key="goal_compare_date",
                )
                actual_basis = st.radio(
                    "실적 기준(양품)",
                    ["양품수량", "샘플제외 양품수량"],
                    index=0,
                    key="goal_actual_basis",
                )
                process_options = [process_display(code) for code in PROCESS_KEYS]
                selected_process_labels = st.multiselect(
                    "공정",
                    options=process_options,
                    default=process_options,
                    key="goal_processes",
                )
                selected_processes = [
                    code
                    for code in PROCESS_KEYS
                    if process_display(code) in selected_process_labels
                ]
                view_mode_label = st.radio(
                    "보기 모드",
                    ["일일(Daily)", "월누적(MTD)", "연누적(YTD)"],
                    index=0,
                    key="goal_view_mode",
                )
                recovery_basis = st.radio(
                    "회복 필요 계산 기준",
                    ["경과일수 기준", "실제근무일수 기준"],
                    index=0,
                    key="goal_recovery_basis",
                )
                red_threshold = st.slider(
                    "진도율 경고 기준(RED)",
                    min_value=0.8,
                    max_value=1.0,
                    value=0.95,
                    step=0.01,
                    key="goal_red_threshold",
                )
                yellow_threshold = st.slider(
                    "진도율 주의 기준(YELLOW)",
                    min_value=red_threshold,
                    max_value=1.2,
                    value=1.0,
                    step=0.01,
                    key="goal_yellow_threshold",
                )
                show_trend = st.checkbox("최근 7일 추이 보기", value=False, key="goal_trend")

            view_mode = (
                "Daily"
                if "Daily" in view_mode_label
                else "MTD"
                if "MTD" in view_mode_label
                else "YTD"
            )

            selected_date = pd.Timestamp(selected_date)
            compare_date = pd.Timestamp(compare_date)
            actual_col = "실적_양품"
            if actual_basis == "샘플제외 양품수량":
                if "실적_샘플제외" in daily_actuals.columns:
                    actual_col = "실적_샘플제외"
                else:
                    st.info("샘플제외 양품수량 컬럼이 없어 양품수량 기준으로 표시합니다.")
            target_mode = "기준자료 근무일수 목표"

            header_placeholder.markdown(
                (
                    "### 일일 생산 현황 보고 (S관/3공장) "
                    f"<span style='font-size:0.9rem; color:#666;'>"
                    f"보고 기준일: {selected_date:%Y-%m-%d}</span>"
                ),
                unsafe_allow_html=True,
            )

            if not selected_processes:
                st.info("선택된 공정이 없어 전체 공정을 표시합니다.")
                selected_processes = PROCESS_KEYS

            summary_actual = 0
            summary_target_ref = 0
            summary_target_actual = 0
            summary_rate_ref = 0
            summary_rate_actual = 0
            summary_prod = 0
            summary_yield = 0
            summary_workdays = 0
            delta_actual = None
            delta_yield = None
            yoy_actual = None
            yoy_rate = None
            yoy_yield = None

            daily_kpi_table = compute_daily_table(
                daily_actuals,
                targets,
                workdays,
                selected_date,
                compare_date,
                selected_processes,
                target_mode,
                "Daily",
                actual_col=actual_col,
            )
            mtd_kpi_table = compute_daily_table(
                daily_actuals,
                targets,
                workdays,
                selected_date,
                compare_date,
                selected_processes,
                target_mode,
                "MTD",
                actual_col=actual_col,
            )

            if view_mode == "Daily":
                table = daily_kpi_table.copy()
                summary_actual = table["실적(선택일)"].sum()
                summary_target_ref = table["목표"].sum()
                summary_target_actual = summary_target_ref
                summary_rate_ref = summary_actual / summary_target_ref if summary_target_ref else 0
                summary_rate_actual = summary_rate_ref
                summary_prod = table["생산수량"].sum()
                summary_yield = summary_actual / summary_prod if summary_prod else 0
                summary_workdays = 0
                if not daily_actuals[
                    (daily_actuals["일자"] == selected_date)
                    & (daily_actuals["공정코드"].isin(selected_processes))
                    & (daily_actuals["실제근무"])
                ].empty:
                    summary_workdays = 1
                prev_actual = table["실적(전일)"].sum()
                prev_prod = daily_actuals[
                    (daily_actuals["일자"] == compare_date)
                    & (daily_actuals["공정코드"].isin(selected_processes))
                ]["생산수량"].sum()
                prev_yield = prev_actual / prev_prod if prev_prod else 0
                delta_actual = summary_actual - prev_actual
                delta_yield = summary_yield - prev_yield

                if selected_date.year == 2026:
                    prev_year_date = selected_date - pd.DateOffset(years=1)
                    prev_year = daily_actuals[
                        (daily_actuals["일자"] == prev_year_date)
                        & (daily_actuals["공정코드"].isin(selected_processes))
                    ]
                    prev_year_actual = prev_year[actual_col].sum()
                    prev_year_prod = prev_year["생산수량"].sum()
                    prev_year_yield = prev_year_actual / prev_year_prod if prev_year_prod else None
                    yoy_actual = summary_actual - prev_year_actual
                    yoy_rate = summary_actual / prev_year_actual - 1 if prev_year_actual else None
                    yoy_yield = (
                        summary_yield - prev_year_yield if prev_year_yield is not None else None
                    )

            elif view_mode == "MTD":
                table = mtd_kpi_table.copy()
                summary_actual = table["실적"].sum()
                summary_prod = table["생산수량"].sum()
                summary_yield = summary_actual / summary_prod if summary_prod else 0
                summary_target_ref = table["월목표_기준자료"].sum()
                summary_target_actual = table["월목표_실제근무"].sum()
                summary_rate_ref = summary_actual / summary_target_ref if summary_target_ref else 0
                summary_rate_actual = summary_actual / summary_target_actual if summary_target_actual else 0
                summary_workdays = daily_actuals[
                    (daily_actuals["일자"] >= selected_date.replace(day=1))
                    & (daily_actuals["일자"] <= selected_date)
                    & (daily_actuals["공정코드"].isin(selected_processes))
                    & (daily_actuals["실제근무"])
                ]["일자"].nunique()

                if selected_date.year == 2026:
                    prev_year_table = compute_daily_table(
                        daily_actuals,
                        targets,
                        workdays,
                        selected_date - pd.DateOffset(years=1),
                        compare_date - pd.DateOffset(years=1),
                        selected_processes,
                        target_mode,
                        "MTD",
                        actual_col=actual_col,
                    )
                    prev_year_actual = prev_year_table["실적"].sum()
                    prev_year_prod = prev_year_table["생산수량"].sum()
                    prev_year_yield = prev_year_actual / prev_year_prod if prev_year_prod else None
                    yoy_actual = summary_actual - prev_year_actual
                    yoy_rate = summary_actual / prev_year_actual - 1 if prev_year_actual else None
                    yoy_yield = (
                        summary_yield - prev_year_yield if prev_year_yield is not None else None
                    )

            else:
                grouped, _ = compute_monthly_kpis(
                    actuals,
                    targets,
                    workdays,
                    selected_date.year,
                    selected_date.month,
                    selected_processes,
                    "연간 누적(YTD)",
                    actual_col=actual_col,
                )
                process_frame = pd.DataFrame({"공정코드": selected_processes})
                grouped = process_frame.merge(grouped, on="공정코드", how="left").fillna(0)
                grouped["공정"] = grouped["공정코드"].apply(process_display)
                table = grouped.copy()
                summary_actual = grouped["실적"].sum()
                summary_prod = grouped["생산수량"].sum()
                summary_yield = summary_actual / summary_prod if summary_prod else 0
                summary_workdays = grouped["실제근무일수"].sum()
                summary_target_ref = grouped["월목표_기준자료"].sum()
                summary_target_actual = grouped["월목표_실제근무"].sum()
                summary_rate_ref = summary_actual / summary_target_ref if summary_target_ref else 0
                summary_rate_actual = summary_actual / summary_target_actual if summary_target_actual else 0

                if selected_date.year == 2026:
                    prev_grouped, _ = compute_monthly_kpis(
                        actuals,
                        targets,
                        workdays,
                        2025,
                        selected_date.month,
                        selected_processes,
                        "연간 누적(YTD)",
                        actual_col=actual_col,
                    )
                    prev_grouped = process_frame.merge(prev_grouped, on="공정코드", how="left").fillna(0)
                    prev_actual = prev_grouped["실적"].sum()
                    prev_prod = prev_grouped["생산수량"].sum()
                    prev_yield = prev_actual / prev_prod if prev_prod else None
                    yoy_actual = summary_actual - prev_actual
                    yoy_rate = summary_actual / prev_actual - 1 if prev_actual else None
                    yoy_yield = summary_yield - prev_yield if prev_yield is not None else None

            if "수율_원칙" in table.columns and "수율(원칙)" not in table.columns:
                table["수율(원칙)"] = table["수율_원칙"]
                table["수율(보정)"] = table["수율_보정"]
            if "이상원인" not in table.columns:
                table["이상원인"] = table["수율이상여부"].apply(
                    lambda value: "분모불일치" if str(value).strip() else ""
                )

            daily_actual_sum = daily_kpi_table["실적(선택일)"].sum()
            daily_target_sum = daily_kpi_table["목표"].sum()
            daily_rate = daily_actual_sum / daily_target_sum if daily_target_sum else 0
            daily_prod_sum = daily_kpi_table["생산수량"].sum()
            daily_yield = daily_actual_sum / daily_prod_sum if daily_prod_sum else 0
            anomaly_count = (daily_kpi_table["수율이상여부"] == "Y").sum()

            mtd_actual_sum = mtd_kpi_table["실적"].sum()
            mtd_target_ref_sum = mtd_kpi_table["월목표_기준자료"].sum()
            mtd_target_actual_sum = mtd_kpi_table["월목표_실제근무"].sum()
            mtd_rate_ref = (
                mtd_actual_sum / mtd_target_ref_sum if mtd_target_ref_sum else 0
            )
            mtd_rate_actual = (
                mtd_actual_sum / mtd_target_actual_sum if mtd_target_actual_sum else 0
            )

            month_start = selected_date.replace(day=1)
            month_actual_base = daily_actuals[
                (daily_actuals["일자"] >= month_start)
                & (daily_actuals["일자"] <= selected_date)
                & (daily_actuals["공정코드"] == "[80]")
            ]
            month_actual_value = month_actual_base[actual_col].sum()
            month_workdays = month_actual_base[
                month_actual_base["실제근무"]
            ]["일자"].nunique()
            month_targets = targets[
                (targets["연"] == selected_date.year)
                & (targets["월"] == selected_date.month)
                & (targets["공정코드"] == "[80]")
            ]
            month_target_daily = (
                month_targets["일일_생산목표량"].sum() if not month_targets.empty else 0
            )
            month_workdays_value = workdays[
                (workdays["연"] == selected_date.year)
                & (workdays["월"] == selected_date.month)
            ]
            month_workdays_value = (
                month_workdays_value["기준자료근무일수"].iloc[0]
                if not month_workdays_value.empty
                else 0
            )
            month_target_value = month_target_daily * month_workdays_value
            month_rate_value = (
                month_actual_value / month_target_value if month_target_value else 0
            )

            month_label = f"{selected_date.month}월"
            kpi_cols = st.columns(4)
            kpi_cols[0].metric(f"{month_label}목표", f"{int(month_target_value):,}")
            kpi_cols[1].metric(f"{month_label}실적", f"{int(month_actual_value):,}")
            kpi_cols[2].metric("실적달성율", f"{month_rate_value * 100:.2f}%")
            kpi_cols[3].metric("실근무일", f"{int(month_workdays):,}")

            color_clear_note = None
            if (
                not actuals_raw.empty
                and "신규분류요약" in actuals_raw.columns
                and "생산일자" in actuals_raw.columns
                and "공장" in actuals_raw.columns
                and "공정코드" in actuals_raw.columns
            ):
                color_source = actuals_raw.copy()
                color_source["공장"] = color_source["공장"].astype(str).str.strip()
                color_source = color_source[color_source["공장"] == "S관(3공장)"]
                color_source["공정코드"] = color_source["공정코드"].apply(
                    normalize_process_code
                )
                color_source = color_source[color_source["공정코드"] == "[80]"]
                color_source["생산일자"] = pd.to_datetime(
                    color_source["생산일자"], errors="coerce"
                )
                color_source = color_source[
                    (color_source["생산일자"] >= month_start)
                    & (color_source["생산일자"] <= selected_date)
                ]
                actual_raw_col = "양품수량"
                if actual_basis == "샘플제외 양품수량":
                    sample_col = find_first_column(
                        color_source.columns, SAMPLE_EXCL_COL_CANDIDATES
                    )
                    if sample_col:
                        actual_raw_col = sample_col
                color_source[actual_raw_col] = normalize_numeric(
                    color_source[actual_raw_col]
                )
                color_mask = color_source["신규분류요약"].astype(str).str.contains(
                    "color", case=False, na=False
                )
                color_sum = color_source.loc[color_mask, actual_raw_col].sum()
                clear_sum = color_source.loc[~color_mask, actual_raw_col].sum()
                color_clear_note = f"Color {int(color_sum):,} · Clear {int(clear_sum):,}"

            if color_clear_note:
                with kpi_cols[1]:
                    st.caption(color_clear_note)

            if delta_actual is not None or delta_yield is not None:
                st.caption(
                    f"전일 대비 실적 {int(delta_actual):,}, 수율 {delta_yield * 100:.2f}p"
                )
            if yoy_actual is not None and yoy_rate is not None:
                st.caption(
                    f"전년 대비 실적 {int(yoy_actual):,} ({yoy_rate * 100:.1f}%), 수율 {yoy_yield * 100:.2f}p"
                )

            if view_mode == "MTD":
                st.caption("MTD는 선택월 1일~기준일까지 누적입니다.")

            if view_mode in {"MTD", "YTD"}:
                table["갭_기준근무"] = table["월목표_기준자료"] - table["실적"]
                table["갭_실제근무"] = table["월목표_실제근무"] - table["실적"]
                if view_mode == "MTD":
                    elapsed_days = table["기준자료근무일수"].clip(
                        upper=int(selected_date.day)
                    )
                else:
                    elapsed_days = table["실제근무일수"]
                if recovery_basis == "경과일수 기준":
                    used_base_days = elapsed_days
                else:
                    used_base_days = table["실제근무일수"]
                table["남은근무일수_기준근무"] = (
                    table["기준자료근무일수"] - used_base_days
                ).clip(lower=0)
                table["남은근무일수_실제근무"] = (
                    table["기준자료근무일수"] - table["실제근무일수"]
                ).clip(lower=0)
                table["회복 필요 일평균(기준근무)"] = (
                    table["갭_기준근무"].clip(lower=0)
                    / table["남은근무일수_기준근무"].replace(0, 1)
                )
                table["회복 필요 일평균(실제근무)"] = (
                    table["갭_실제근무"].clip(lower=0)
                    / table["남은근무일수_실제근무"].replace(0, 1)
                )

                def status_for(rate):
                    if pd.isna(rate):
                        return ""
                    if rate >= yellow_threshold:
                        return "Green"
                    if rate >= red_threshold:
                        return "Yellow"
                    return "Red"

                table["상태"] = table["달성율_기준자료"].apply(status_for)

            target_map = targets[
                (targets["연"] == selected_date.year)
                & (targets["월"] == selected_date.month)
                & (targets["공정코드"].isin(PROCESS_KEYS))
            ].groupby("공정코드")["일일_생산목표량"].sum()
            monthly_actual_col = (
                "실적_양품" if "실적_양품" in daily_actuals.columns else actual_col
            )
            month_actual_map = daily_actuals[
                (daily_actuals["일자"] >= month_start)
                & (daily_actuals["일자"] <= selected_date)
                & (daily_actuals["공정코드"].isin(PROCESS_KEYS))
            ].groupby("공정코드")[monthly_actual_col].sum()
            prev_year_date = selected_date - pd.DateOffset(years=1)
            prev_year_start = prev_year_date.replace(day=1)
            prev_year_end = prev_year_start + pd.offsets.MonthEnd(0)
            prev_year_cutoff = min(
                prev_year_end,
                prev_year_start + pd.Timedelta(days=selected_date.day - 1),
            )
            prev_year_actual_map = daily_actuals[
                (daily_actuals["일자"] >= prev_year_start)
                & (daily_actuals["일자"] <= prev_year_cutoff)
                & (daily_actuals["공정코드"].isin(PROCESS_KEYS))
            ].groupby("공정코드")[monthly_actual_col].sum()

            prev_month_date = selected_date - pd.DateOffset(months=1)
            prev_month_start = prev_month_date.replace(day=1)
            prev_month_end = prev_month_start + pd.offsets.MonthEnd(0)
            prev_month_cutoff = min(
                prev_month_end,
                prev_month_start + pd.Timedelta(days=selected_date.day - 1),
            )
            prev_month_actual_map = daily_actuals[
                (daily_actuals["일자"] >= prev_month_start)
                & (daily_actuals["일자"] <= prev_month_cutoff)
                & (daily_actuals["공정코드"].isin(PROCESS_KEYS))
            ].groupby("공정코드")[monthly_actual_col].sum()

            monthly_rows = []
            workdays_map = daily_actuals[
                (daily_actuals["일자"] >= month_start)
                & (daily_actuals["일자"] <= selected_date)
                & (daily_actuals["공정코드"].isin(PROCESS_KEYS))
                & (daily_actuals["실제근무"])
            ].groupby("공정코드")["일자"].nunique()
            prev_year_workdays_map = daily_actuals[
                (daily_actuals["일자"] >= prev_year_start)
                & (daily_actuals["일자"] <= prev_year_cutoff)
                & (daily_actuals["공정코드"].isin(PROCESS_KEYS))
                & (daily_actuals["실제근무"])
            ].groupby("공정코드")["일자"].nunique()
            prev_month_workdays_map = daily_actuals[
                (daily_actuals["일자"] >= prev_month_start)
                & (daily_actuals["일자"] <= prev_month_cutoff)
                & (daily_actuals["공정코드"].isin(PROCESS_KEYS))
                & (daily_actuals["실제근무"])
            ].groupby("공정코드")["일자"].nunique()
            for code in PROCESS_KEYS:
                daily_target = target_map.get(code, 0)
                workdays_count = workdays_map.get(code, 0)
                month_target = daily_target * workdays_count
                month_actual = month_actual_map.get(code, 0)
                prev_year_actual = prev_year_actual_map.get(code, 0)
                prev_month_actual = prev_month_actual_map.get(code, 0)
                rate = month_actual / month_target if month_target else pd.NA
                avg_target = month_target / workdays_count if workdays_count else pd.NA
                avg_actual = month_actual / workdays_count if workdays_count else pd.NA
                prev_year_days = prev_year_workdays_map.get(code, 0)
                prev_month_days = prev_month_workdays_map.get(code, 0)
                prev_year_avg = (
                    prev_year_actual / prev_year_days if prev_year_days else pd.NA
                )
                prev_month_avg = (
                    prev_month_actual / prev_month_days if prev_month_days else pd.NA
                )
                monthly_rows.append(
                    {
                        ("공정", ""): process_display(code),
                        ("목표", "일"): avg_target,
                        ("목표", "월"): month_target,
                        ("실적", "일"): avg_actual,
                        ("실적", "월"): month_actual,
                        ("달성율", "(실적/목표)"): rate,
                        ("전년동월", "일"): prev_year_avg,
                        ("전년동월", "월"): prev_year_actual,
                        ("전년동월", "증감(월)"): month_actual - prev_year_actual,
                        ("전월", "일"): prev_month_avg,
                        ("전월", "월"): prev_month_actual,
                        ("전월", "증감(월)"): month_actual - prev_month_actual,
                    }
                )

            monthly_table = pd.DataFrame(monthly_rows)
            monthly_table.columns = pd.MultiIndex.from_tuples(monthly_table.columns)
            st.markdown("#### 공정별 목표/실적")

            def format_achievement(value):
                if pd.isna(value):
                    return ""
                if value >= 1:
                    icon = "✅"
                elif value >= 0.9:
                    icon = "⚠️"
                else:
                    icon = "❌"
                return f"{icon} {value * 100:.2f}%"

            def highlight_delta_columns(data):
                styles = pd.DataFrame("", index=data.index, columns=data.columns)
                delta_columns = [
                    ("전년동월", "증감(월)"),
                    ("전월", "증감(월)"),
                ]
                for col in delta_columns:
                    if col not in data.columns:
                        continue
                    for idx, value in data[col].items():
                        if pd.isna(value):
                            continue
                        if value < 0:
                            styles.loc[idx, col] = "color: #D0021B; font-weight: 700;"
                        elif value > 0:
                            styles.loc[idx, col] = "color: #2B6CB0; font-weight: 700;"
                return styles

            monthly_styled = apply_alignment(
                monthly_table.style.format(
                    {
                        ("목표", "일"): "{:,.0f}",
                        ("목표", "월"): "{:,.0f}",
                        ("실적", "일"): "{:,.0f}",
                        ("실적", "월"): "{:,.0f}",
                        ("전년동월", "일"): "{:,.0f}",
                        ("전년동월", "월"): "{:,.0f}",
                        ("전년동월", "증감(월)"): "{:,.0f}",
                        ("전월", "일"): "{:,.0f}",
                        ("전월", "월"): "{:,.0f}",
                        ("전월", "증감(월)"): "{:,.0f}",
                        ("달성율", "(실적/목표)"): format_achievement,
                    },
                    na_rep="",
                ).apply(highlight_delta_columns, axis=None),
                monthly_table,
            )
            st.dataframe(
                monthly_styled,
                width="stretch",
                height=calc_table_height(len(monthly_table), max_height=420),
            )

            yield_process_map = {
                "[10]": "사출조립",
                "[20]": "분리",
                "[45]": "수화/검사",
                "[55]": "접착/멸균",
                "[80]": "누수/규격검사",
            }
            yield_source = daily_actuals[
                (daily_actuals["일자"] >= month_start)
                & (daily_actuals["일자"] <= selected_date)
                & (daily_actuals["공정코드"].isin(yield_process_map.keys()))
            ].copy()
            prev_year_yield_source = daily_actuals[
                (daily_actuals["일자"] >= prev_year_start)
                & (daily_actuals["일자"] <= prev_year_cutoff)
                & (daily_actuals["공정코드"].isin(yield_process_map.keys()))
            ].copy()
            prev_month_yield_source = daily_actuals[
                (daily_actuals["일자"] >= prev_month_start)
                & (daily_actuals["일자"] <= prev_month_cutoff)
                & (daily_actuals["공정코드"].isin(yield_process_map.keys()))
            ].copy()

            yield_grouped = (
                yield_source.groupby("공정코드", dropna=False)
                .agg(실적=(actual_col, "sum"), 생산수량=("생산수량", "sum"))
                .reset_index()
            )
            yield_grouped["수율"] = (
                yield_grouped["실적"]
                / yield_grouped["생산수량"].replace(0, pd.NA)
            )

            prev_year_yield_grouped = (
                prev_year_yield_source.groupby("공정코드", dropna=False)
                .agg(실적=(actual_col, "sum"), 생산수량=("생산수량", "sum"))
                .reset_index()
            )
            prev_year_yield_grouped["수율"] = (
                prev_year_yield_grouped["실적"]
                / prev_year_yield_grouped["생산수량"].replace(0, pd.NA)
            )

            prev_month_yield_grouped = (
                prev_month_yield_source.groupby("공정코드", dropna=False)
                .agg(실적=(actual_col, "sum"), 생산수량=("생산수량", "sum"))
                .reset_index()
            )
            prev_month_yield_grouped["수율"] = (
                prev_month_yield_grouped["실적"]
                / prev_month_yield_grouped["생산수량"].replace(0, pd.NA)
            )

            current_label = f"{selected_date.month}월 수율"
            yield_rows = []
            for code, label in yield_process_map.items():
                current_row = yield_grouped[yield_grouped["공정코드"] == code]
                prev_year_row = prev_year_yield_grouped[
                    prev_year_yield_grouped["공정코드"] == code
                ]
                prev_month_row = prev_month_yield_grouped[
                    prev_month_yield_grouped["공정코드"] == code
                ]
                current_rate = (
                    current_row["수율"].iloc[0] if not current_row.empty else pd.NA
                )
                prev_year_rate = (
                    prev_year_row["수율"].iloc[0] if not prev_year_row.empty else pd.NA
                )
                prev_month_rate = (
                    prev_month_row["수율"].iloc[0] if not prev_month_row.empty else pd.NA
                )
                diff_year = (
                    current_rate - prev_year_rate
                    if pd.notna(current_rate) and pd.notna(prev_year_rate)
                    else pd.NA
                )
                diff_month = (
                    current_rate - prev_month_rate
                    if pd.notna(current_rate) and pd.notna(prev_month_rate)
                    else pd.NA
                )
                yield_rows.append(
                    {
                        "공정": label,
                        "전년동월 수율": prev_year_rate,
                        "전월 수율": prev_month_rate,
                        current_label: current_rate,
                        "증감(전년대비)": diff_year,
                        "증감(전월대비)": diff_month,
                    }
                )
            current_vals = [
                row[current_label] for row in yield_rows if pd.notna(row[current_label])
            ]
            prev_year_vals = [
                row["전년동월 수율"]
                for row in yield_rows
                if pd.notna(row["전년동월 수율"])
            ]
            prev_month_vals = [
                row["전월 수율"] for row in yield_rows if pd.notna(row["전월 수율"])
            ]
            total_yield = 1
            for val in current_vals:
                total_yield *= val
            total_prev_year = 1
            for val in prev_year_vals:
                total_prev_year *= val
            total_prev_month = 1
            for val in prev_month_vals:
                total_prev_month *= val
            yield_rows.append(
                {
                    "공정": "종합수율",
                    "전년동월 수율": total_prev_year if prev_year_vals else pd.NA,
                    "전월 수율": total_prev_month if prev_month_vals else pd.NA,
                    current_label: total_yield if current_vals else pd.NA,
                    "증감(전년대비)": (
                        total_yield - total_prev_year
                        if current_vals and prev_year_vals
                        else pd.NA
                    ),
                    "증감(전월대비)": (
                        total_yield - total_prev_month
                        if current_vals and prev_month_vals
                        else pd.NA
                    ),
                }
            )
            yield_table = pd.DataFrame(yield_rows)

            def highlight_yield(data):
                styles = pd.DataFrame("", index=data.index, columns=data.columns)
                for col in ["증감(전년대비)", "증감(전월대비)"]:
                    if col not in data.columns:
                        continue
                    for idx, value in data[col].items():
                        if pd.isna(value):
                            continue
                        if value < 0:
                            styles.loc[idx, col] = "color: #D0021B; font-weight: 700;"
                        elif value > 0:
                            styles.loc[idx, col] = "color: #2B6CB0; font-weight: 700;"
                return styles

            display_yield = yield_table
            st.markdown("#### 공정별 수율")
            yield_styled = apply_alignment(
                display_yield.style.format(
                    {
                        "전년동월 수율": "{:.2%}",
                        "전월 수율": "{:.2%}",
                        current_label: "{:.2%}",
                        "증감(전년대비)": "{:+.2%}",
                        "증감(전월대비)": "{:+.2%}",
                    },
                    na_rep="",
                ).apply(highlight_yield, axis=None),
                display_yield,
            )
            st.dataframe(
                yield_styled,
                width="stretch",
                height=calc_table_height(len(display_yield), max_height=420),
            )

            daily_month = daily_actuals[
                (daily_actuals["일자"] >= month_start)
                & (daily_actuals["일자"] <= selected_date)
                & (daily_actuals["공정코드"] == "[80]")
            ].copy()
            if not daily_month.empty:
                daily_sum = (
                    daily_month.groupby("일자", dropna=False)[actual_col]
                    .sum()
                    .reset_index()
                    .rename(columns={actual_col: "실적"})
                )
                daily_sum["목표(일)"] = target_map.get("[80]", 0)
                daily_sum["차이"] = daily_sum["실적"] - daily_sum["목표(일)"]
                daily_sum["달성율"] = (
                    daily_sum["실적"] / daily_sum["목표(일)"].replace(0, pd.NA)
                )
                daily_sum["일자"] = daily_sum["일자"].dt.strftime("%Y-%m-%d")
            else:
                daily_sum = pd.DataFrame(
                    columns=["일자", "목표(일)", "실적", "차이", "달성율"]
                )

            st.markdown("#### 금월 일별 실적")

            def highlight_daily_diff(data):
                styles = pd.DataFrame("", index=data.index, columns=data.columns)
                if "차이" in data.columns:
                    for idx, value in data["차이"].items():
                        if pd.isna(value):
                            continue
                        if value < 0:
                            styles.loc[idx, "차이"] = "color: #D0021B; font-weight: 700;"
                        elif value > 0:
                            styles.loc[idx, "차이"] = "color: #2B6CB0; font-weight: 700;"
                return styles

            daily_styled = apply_alignment(
                daily_sum.style.format(
                    {
                        "목표(일)": "{:,.0f}",
                        "실적": "{:,.0f}",
                        "차이": "{:,.0f}",
                        "달성율": format_achievement,
                    },
                    na_rep="",
                ).apply(highlight_daily_diff, axis=None),
                daily_sum,
            )
            st.dataframe(
                daily_styled,
                width="stretch",
                height=calc_table_height(len(daily_sum), max_height=520),
            )

            if view_mode == "Daily":
                rate_data = table[["공정", "달성율(선택일)"]].copy()
                rate_data = rate_data.rename(columns={"달성율(선택일)": "달성율"})
                rate_data["구분"] = "선택일"
                gap_data = table[["공정", "목표", "실적(선택일)"]].copy()
                gap_data["갭"] = (gap_data["목표"] - gap_data["실적(선택일)"]).clip(lower=0)
            else:
                rate_data = table.melt(
                    id_vars=["공정"],
                    value_vars=["달성율_기준자료", "달성율_실제근무"],
                    var_name="구분",
                    value_name="달성율",
                )
                gap_data = table[["공정", "갭_기준근무"]].copy()
                gap_data = gap_data.rename(columns={"갭_기준근무": "갭"})

            rate_chart = (
                alt.Chart(rate_data)
                .mark_bar()
                .encode(
                    x=alt.X("공정:N"),
                    y=alt.Y("달성율:Q", axis=alt.Axis(format="%")),
                    color=alt.Color("구분:N"),
                    tooltip=["공정", "구분", alt.Tooltip("달성율:Q", format=".2%")],
                )
                .properties(height=260)
            )
            rate_rule = alt.Chart(
                pd.DataFrame({"threshold": [red_threshold]})
            ).mark_rule(color="#D0021B", strokeDash=[6, 4]).encode(y="threshold:Q")
            st.altair_chart(rate_chart + rate_rule, use_container_width=True)

            gap_chart = (
                alt.Chart(gap_data)
                .mark_bar(color="#D0021B")
                .encode(
                    x=alt.X("공정:N"),
                    y=alt.Y("갭:Q", title="갭(목표-실적)"),
                    tooltip=["공정", alt.Tooltip("갭:Q", format=",")],
                )
                .properties(height=220)
            )
            st.altair_chart(gap_chart, use_container_width=True)

            if view_mode == "Daily" and show_trend:
                start_trend = selected_date - pd.Timedelta(days=6)
                trend_dates = pd.date_range(start_trend, selected_date, freq="D")
                trend = daily_actuals[
                    (daily_actuals["일자"].isin(trend_dates))
                    & (daily_actuals["공정코드"].isin(selected_processes))
                ]
                trend_summary = (
                    trend.groupby("일자", dropna=False)
                    .agg(실적=(actual_col, "sum"), 생산수량=("생산수량", "sum"))
                    .reset_index()
                )
                trend_summary["수율"] = (
                    trend_summary["실적"] / trend_summary["생산수량"].replace(0, pd.NA)
                )
                target_map = targets[
                    (targets["연"] == selected_date.year)
                    & (targets["월"] == selected_date.month)
                    & (targets["공정코드"].isin(selected_processes))
                ]["일일_생산목표량"].sum()
                trend_summary["목표"] = target_map
                trend_melt = trend_summary.melt(
                    id_vars=["일자"],
                    value_vars=["실적", "목표", "수율"],
                    var_name="구분",
                    value_name="수량",
                )
                trend_chart = (
                    alt.Chart(trend_melt)
                    .mark_line(point=True)
                    .encode(
                        x=alt.X("일자:T", title="일자"),
                        y=alt.Y("수량:Q", title="수량"),
                        color=alt.Color("구분:N"),
                        tooltip=["일자", "구분", "수량"],
                    )
                    .properties(height=240)
                )
                st.altair_chart(trend_chart, use_container_width=True)

            if view_mode == "Daily":
                period_start = selected_date
                period_end = selected_date
            elif view_mode == "MTD":
                period_start = selected_date.replace(day=1)
                period_end = selected_date
            else:
                period_start = pd.Timestamp(year=selected_date.year, month=1, day=1)
                period_end = selected_date

            anomaly_detail = daily_actuals[
                (daily_actuals["일자"] >= period_start)
                & (daily_actuals["일자"] <= period_end)
                & (daily_actuals["공정코드"].isin(selected_processes))
            ].copy()
            if not anomaly_detail.empty:
                anomaly_detail["총출력"] = (
                    anomaly_detail["실적_양품"]
                    + anomaly_detail["불량수량"]
                    + anomaly_detail["샘플수량"]
                )
                anomaly_detail["수율이상여부"] = (
                    anomaly_detail["생산수량"] < anomaly_detail["총출력"]
                )
                denom = anomaly_detail[["생산수량", "총출력"]].max(axis=1).replace(0, pd.NA)
                anomaly_detail["수율(원칙)"] = (
                    anomaly_detail[actual_col] / anomaly_detail["생산수량"].replace(0, pd.NA)
                )
                anomaly_detail["수율(보정)"] = anomaly_detail[actual_col] / denom
                anomaly_detail["공정"] = anomaly_detail["공정코드"].apply(process_display)
                anomaly_detail = anomaly_detail[anomaly_detail["수율이상여부"]]

            priority_source = table.copy() if view_mode in {"MTD", "YTD"} else mtd_kpi_table.copy()
            if "갭_기준근무" not in priority_source.columns:
                priority_source["갭_기준근무"] = (
                    priority_source["월목표_기준자료"] - priority_source["실적"]
                )
                priority_source["갭_실제근무"] = (
                    priority_source["월목표_실제근무"] - priority_source["실적"]
                )
                if view_mode == "MTD":
                    elapsed_days = priority_source["기준자료근무일수"].clip(
                        upper=int(selected_date.day)
                    )
                else:
                    elapsed_days = priority_source["실제근무일수"]
                if recovery_basis == "경과일수 기준":
                    used_base_days = elapsed_days
                else:
                    used_base_days = priority_source["실제근무일수"]
                priority_source["남은근무일수_기준근무"] = (
                    priority_source["기준자료근무일수"] - used_base_days
                ).clip(lower=0)
                priority_source["남은근무일수_실제근무"] = (
                    priority_source["기준자료근무일수"] - priority_source["실제근무일수"]
                ).clip(lower=0)
                priority_source["회복 필요 일평균(기준근무)"] = (
                    priority_source["갭_기준근무"].clip(lower=0)
                    / priority_source["남은근무일수_기준근무"].replace(0, 1)
                )
                priority_source["회복 필요 일평균(실제근무)"] = (
                    priority_source["갭_실제근무"].clip(lower=0)
                    / priority_source["남은근무일수_실제근무"].replace(0, 1)
                )

            st.markdown("#### 지적 우선순위 Top 3")
            top_gap = (
                priority_source.sort_values("갭_기준근무", ascending=False)
                .head(3)
                .loc[:, ["공정", "갭_기준근무"]]
            )
            top_recovery = (
                priority_source.sort_values("회복 필요 일평균(기준근무)", ascending=False)
                .head(3)
                .loc[:, ["공정", "회복 필요 일평균(기준근무)"]]
            )
            top_drop = (
                daily_kpi_table.sort_values("실적 증감")
                .head(3)
                .loc[:, ["공정", "실적 증감"]]
            )

            top_cols = st.columns(3)
            top_cols[0].markdown(
                "목표대비 부족량\n"
                + "\n".join(
                    f"- {row['공정']}: {row['갭_기준근무']:,.0f}"
                    for _, row in top_gap.iterrows()
                )
            )
            top_cols[1].markdown(
                "회복 필요 일평균\n"
                + "\n".join(
                    f"- {row['공정']}: {row['회복 필요 일평균(기준근무)']:,.0f}"
                    for _, row in top_recovery.iterrows()
                )
            )
            top_cols[2].markdown(
                "전일 대비 하락\n"
                + "\n".join(
                    f"- {row['공정']}: {row['실적 증감']:,.0f}"
                    for _, row in top_drop.iterrows()
                )
            )

            st.markdown("#### 한 줄 코멘트")
            comment_lines = []
            for _, row in daily_kpi_table.iterrows():
                process_name = row["공정"]
                mtd_row = priority_source[priority_source["공정"] == process_name]
                gap_val = mtd_row["갭_기준근무"].iloc[0] if not mtd_row.empty else pd.NA
                recovery_val = (
                    mtd_row["회복 필요 일평균(기준근무)"].iloc[0] if not mtd_row.empty else pd.NA
                )
                rate_val = row.get("달성율(선택일)", pd.NA)
                rate_text = f"{rate_val * 100:.1f}%" if pd.notna(rate_val) else "-"
                delta_rate = row.get("달성율 증감", pd.NA)
                delta_rate_text = (
                    f"{delta_rate * 100:.1f}p" if pd.notna(delta_rate) else "-"
                )
                gap_text = f"{gap_val:,.0f}" if pd.notna(gap_val) else "-"
                recovery_text = (
                    f"{recovery_val:,.0f}" if pd.notna(recovery_val) else "-"
                )
                comment_lines.append(
                    f"- {process_name}: 금일 달성율 {rate_text}"
                    f", 전일 대비 {delta_rate_text}. "
                    f"월목표(기준근무) 대비 갭 {gap_text}, "
                    f"남은 근무일 기준 일평균 {recovery_text} 필요."
                )
            if anomaly_detail is not None and not anomaly_detail.empty:
                comment_lines.append(
                    f"- 수율 이상(경고) 공정 존재: {anomaly_detail['공정'].nunique()}개 공정, "
                    f"{len(anomaly_detail):,}건 (분모불일치)."
                )
            st.markdown("\n".join(comment_lines) if comment_lines else "- 코멘트 생성 데이터 없음.")

            with st.expander("수율 이상 상세"):
                if anomaly_detail is None or anomaly_detail.empty:
                    st.caption("수율 이상 케이스가 없습니다.")
                else:
                    detail_cols = [
                        "일자",
                        "공정",
                        "생산수량",
                        "실적_양품",
                        "불량수량",
                        "샘플수량",
                        "총출력",
                        "수율(원칙)",
                        "수율(보정)",
                    ]
                    detail_view = anomaly_detail[detail_cols].copy()
                    detail_view["일자"] = detail_view["일자"].dt.strftime("%Y-%m-%d")
                    detail_format = {}
                    for col in detail_view.columns:
                        if pd.api.types.is_numeric_dtype(detail_view[col]):
                            detail_format[col] = "{:,.2f}" if "수율" in col else "{:,.0f}"
                    detail_styled = apply_alignment(
                        detail_view.style.format(detail_format, na_rep=""),
                        detail_view,
                    )
                    st.dataframe(
                        detail_styled,
                        width="stretch",
                        height=calc_table_height(len(detail_view), max_height=320),
                    )

            with st.expander("원자료 집계 검증(공정별 생산수량 vs 양품+불량+샘플)"):
                validation = daily_actuals[
                    (daily_actuals["일자"] >= period_start)
                    & (daily_actuals["일자"] <= period_end)
                    & (daily_actuals["공정코드"].isin(selected_processes))
                ].copy()
                if validation.empty:
                    st.caption("검증할 데이터가 없습니다.")
                else:
                    validation["총출력"] = (
                        validation["실적_양품"]
                        + validation["불량수량"]
                        + validation["샘플수량"]
                    )
                    validation = (
                        validation.groupby("공정코드", dropna=False)
                        .agg(
                            생산수량=("생산수량", "sum"),
                            총출력=("총출력", "sum"),
                        )
                        .reset_index()
                    )
                    validation["차이"] = validation["생산수량"] - validation["총출력"]
                    validation["공정"] = validation["공정코드"].apply(process_display)
                    validation_view = validation[["공정", "생산수량", "총출력", "차이"]]
                    validation_format = {
                        "생산수량": "{:,.0f}",
                        "총출력": "{:,.0f}",
                        "차이": "{:,.0f}",
                    }
                    validation_styled = apply_alignment(
                        validation_view.style.format(validation_format, na_rep=""),
                        validation_view,
                    )
                    st.dataframe(
                        validation_styled,
                        width="stretch",
                        height=calc_table_height(len(validation), max_height=280),
                    )

            summary_df = pd.DataFrame(
                [
                    ("실적 기준", actual_basis),
                    ("금월목표", month_target_value),
                    ("금월실적", month_actual_value),
                    ("실적달성율", month_rate_value),
                    ("실근무일", month_workdays),
                    ("보기 모드", view_mode_label),
                ],
                columns=["항목", "값"],
            )
            report_bytes, report_error = build_excel_report(
                "일일 생산 현황 보고 (S관/3공장)",
                selected_date.strftime("%Y-%m-%d"),
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                summary_df,
                monthly_table,
            )
            if report_bytes is None:
                st.warning("Excel 파일을 생성할 수 없습니다. CSV로 다운로드합니다.")
                csv_bytes = monthly_table.to_csv(index=False, encoding="utf-8-sig").encode(
                    "utf-8-sig"
                )
                st.download_button(
                    "일일 생산 현황 CSV 다운로드",
                    data=csv_bytes,
                    file_name=f"일일_생산_현황_{selected_date.strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                )
            else:
                st.download_button(
                    "일일 생산 현황 보고 Excel 다운로드",
                    data=report_bytes,
                    file_name=f"일일_생산_현황_{selected_date.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            empty_png = (
                b"\x89PNG\r\n\x1a\n"
                b"\x00\x00\x00\rIHDR\x00\x00\x00\x01"
                b"\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89"
                b"\x00\x00\x00\nIDATx\x9cc`\x00\x00\x00\x02\x00\x01"
                b"\xe2!\xbc3\x00\x00\x00\x00IEND\xaeB`\x82"
            )
            st.download_button(
                "일일 생산 현황 이미지(PNG) 다운로드",
                data=empty_png,
                file_name=f"일일_생산_현황_{selected_date.strftime('%Y%m%d')}.png",
                mime="image/png",
            )

            # 가정/확인 포인트:
            # 1) 일일_생산목표량은 일당 목표로 가정.
            # 2) 실제근무일수는 (양품수량 > 0 또는 생산수량 > 0) 기준.
            # 3) 공정코드 값은 실적/목표 파일에 동일하게 존재한다고 가정.


if __name__ == "__main__":
    main()

